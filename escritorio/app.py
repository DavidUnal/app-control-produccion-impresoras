# desktop/app.py
from __future__ import annotations

import sys
import json
import os
import atexit
import datetime
from functools import partial
from pathlib import Path
from urllib.parse import urlencode

import requests
from PySide6.QtCore import Qt, QDate, QSize, QUrl, QTimer, Signal, QObject , QSettings , QEvent 
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QStackedWidget, QListWidget, QSplitter,
    QHBoxLayout, QVBoxLayout, QFormLayout, QLabel, QLineEdit, QDoubleSpinBox, QSpinBox,
    QMessageBox, QPushButton, QComboBox, QDateEdit, QTabWidget, QListWidget as QList,
    QTableWidgetItem, QHeaderView, QCheckBox, QScrollArea, QTableWidget, QFileDialog,
    QPlainTextEdit, QDialog, QFrame, QAbstractItemView, QStyle, QListWidgetItem,
    QSizePolicy, QInputDialog, QGraphicsDropShadowEffect, QSpacerItem, QToolButton, QMenu ,QGroupBox , 
    QStackedLayout , QColorDialog
)
from PySide6.QtGui import (
    QDesktopServices, QShortcut, QKeySequence, QIcon, QPixmap, QFont, QAction, QColor, QPainter,  QStandardItemModel,
    QStandardItem , QAction
)
from PySide6.QtWebSockets import QWebSocket


# stdout line-buffered (Windows suele agradecerlo)
try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass





# =========================
# Imports locales (paquete / script)
# =========================
try:
    from .ui_theme import apply_theme
except Exception:
    from ui_theme import apply_theme

try:
    # paquete
    from .admin_solicitudes import (
        AdminSolicitudesPage as AdminSolicitudesPendientesPage,
        AdminSolicitudesFrame,
    )
except Exception:
    # script
    from admin_solicitudes import (
        AdminSolicitudesPage as AdminSolicitudesPendientesPage,
        AdminSolicitudesFrame,
    )

try:
    from .auth_google_desktop import signin_with_google  # paquete
except Exception:
    try:
        from auth_google_desktop import signin_with_google  # script
    except Exception:
        signin_with_google = None

def _safe_json(resp: requests.Response):
    try:
        return resp.json()
    except Exception:
        return None
    
# =========================
# Rutas/recursos (DEV + PyInstaller)  ✅ orden correcto
# =========================
def _resource_root() -> Path:
    """
    - En DEV: carpeta raíz del proyecto (padre de /escritorio)
    - En PyInstaller:
        * recursos: sys._MEIPASS
        * config: normalmente al lado del .exe (lo manejamos en _load_api_base_url)
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parents[1]


ROOT_DIR: Path = _resource_root()


def _find_logo_in_repo(root: Path) -> Path | None:
    candidates = [
        "logo marca.png", "logo marca.jpg", "logo marca.jpeg", "logo marca.ico",
        "logo marca.bmp", "logo marca.svg",
        "logo_marca.png", "logo_marca.jpg", "logo_marca.jpeg", "logo_marca.ico",
    ]
    for name in candidates:
        p = root / name
        if p.exists():
            return p
    for ext in (".png", ".jpg", ".jpeg", ".ico", ".bmp", ".svg"):
        p = root / f"logo marca{ext}"
        if p.exists():
            return p
    return None


BG_LOGIN_PATH: Path = ROOT_DIR / "fondo de imagen.jpg"
BG_APP_PATH:   Path = ROOT_DIR / "fondo_app.jpg"
APP_BG_PATH:   Path = BG_APP_PATH

LOGO_PATH: Path = (_find_logo_in_repo(ROOT_DIR) or (ROOT_DIR / "logo marca.ico"))


def _load_api_base_url() -> str:
    """
    1) Lee desktop_config.json en la carpeta del EXE (instalación)
    2) Si no está, intenta en _internal
    3) Si no, usa PUBLIC_BASE_URL
    4) fallback localhost
    """
    cfg_candidates = [
        ROOT_DIR / "desktop_config.json",
        ROOT_DIR / "_internal" / "desktop_config.json",
        ROOT_DIR / "_internal" / "desktop" / "desktop_config.json",
    ]

    for cfg_path in cfg_candidates:
        if cfg_path.exists():
            try:
                data = json.loads(cfg_path.read_text(encoding="utf-8"))
                url = (data.get("API_BASE_URL") or data.get("PUBLIC_BASE_URL") or "").strip()
                if url:
                    return url
            except Exception:
                pass

    return os.getenv("PUBLIC_BASE_URL", "http://127.0.0.1:8000")


def resource_path(*parts: str) -> str:
    """
    Devuelve ruta absoluta a un recurso.
    - En DEV: relativo a la carpeta raíz del proyecto.
    - En PyInstaller: relativo a sys._MEIPASS.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).resolve().parents[1]
    return str(base.joinpath(*parts))


BASE_URL = _load_api_base_url()
API = BASE_URL
TIMEOUT = 800

_OAUTH_LOG_FH = open("desktop_oauth.log", "a", encoding="utf-8")
atexit.register(lambda: _OAUTH_LOG_FH.close())


COMPANY_YELLOW = "#F2C037"
PENDIENTE_BG   = "#E0E0E0"
EN_PROCESO_BG  = COMPANY_YELLOW
FINALIZADA_BG  = "#48a259"
CANCELADA_BG   = "#AE0E16"



def _dbg(msg: str) -> None:
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    try:
        _OAUTH_LOG_FH.write(line + "\n")
        _OAUTH_LOG_FH.flush()
    except Exception:
        pass

# --- Auth helpers HTTP ---
def _auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"} if token else {}

def get_pendientes(base_url: str, token: str):
    r = requests.get(
        f"{base_url}/auth/admin/subusuarios/pendientes",
        headers=_auth_headers(token),
        timeout=10
    )
    r.raise_for_status()
    return r.json()
        
def aprobar_usuario(base_url: str, token: str, uid: int, rol: str):
    r = requests.post(
        f"{base_url}/auth/admin/subusuarios/{uid}/aprobar",
        json={"rol": rol},
        headers=_auth_headers(token),
        timeout=10,
    )
    r.raise_for_status()
    return r.json()

def set_rol(base_url: str, token: str, uid: int, rol: str):
    r = requests.post(
        f"{base_url}/auth/admin/subusuarios/{uid}/set-rol",
        json={"rol": rol},
        headers=_auth_headers(token),
        timeout=10,
    )
    r.raise_for_status()
    return r.json()

def rechazar_usuario(base_url: str, token: str, uid: int):
    try:
        r = requests.post(f"{base_url}/auth/admin/usuarios/{uid}/rechazar",
                          headers=_auth_headers(token), timeout=10)
        if r.status_code < 400:
            return r.json()
    except Exception:
        pass
    return set_rol(base_url, token, uid, "RECHAZADO")



# === Estado de sesión en cliente ===
AUTH = {"token": None}
API_TOKEN: str | None = None
USER_ROLES: list[str] = []

def set_token(tok: str | None):
    AUTH["token"] = tok

def _headers():
    return {"Authorization": f"Bearer {AUTH['token']}"} if AUTH["token"] else {}

# --- Utilidades UI / iconos ---
def _std_icon(widget, sp: QStyle.StandardPixmap) -> QIcon:
    return widget.style().standardIcon(sp)

def _icon_from_assets_or_std(widget, filename: str, fallback_sp: QStyle.StandardPixmap) -> QIcon:
    base = os.path.join(os.path.dirname(__file__), "assets", "icons")
    path = os.path.join(base, filename)
    if os.path.exists(path):
        pm = QPixmap(path)
        if not pm.isNull():
            return QIcon(pm)
    return _std_icon(widget, fallback_sp)


def _signin_with_google_compat(self_widget=None):
    """
    Llama a signin_with_google intentando varias firmas posibles:
    - parent=self (keyword)
    - self (posicional)
    - sin argumentos
    Devuelve el dict con el token o None si falla.
    """
    if not signin_with_google:
        return None

    # 1) con keyword parent=self
    try:
        return signin_with_google(parent=self_widget)
    except TypeError:
        pass

    # 2) con argumento posicional self
    try:
        return signin_with_google(self_widget)
    except TypeError:
        pass

    # 3) sin argumentos
    try:
        return signin_with_google()
    except TypeError:
        return None






def _beautify_exist_table(tbl: QTableWidget):
    tbl.horizontalHeader().setStyleSheet(
        f"""
        QHeaderView::section {{
            background: {COMPANY_YELLOW};
            color: #222;
            font-weight: 600;
            border: 1px solid #e8c35c;
            padding: 6px 8px;
        }}
        """
    )
    tbl.setAlternatingRowColors(True)
    tbl.setStyleSheet(tbl.styleSheet() + " QTableWidget { gridline-color: #E0E0E0; } ")
    hdr = tbl.horizontalHeader()
    hdr.setStretchLastSection(False)
    hdr.setSectionResizeMode(QHeaderView.Interactive)

def _icon_or_fallback(path: str | Path | None, widget: QWidget, fallback_sp: QStyle.StandardPixmap) -> QIcon:
    """QIcon desde archivo si existe; si no, icono estándar."""
    try:
        if path and os.path.exists(path):
            pm = QPixmap(str(path))
            if not pm.isNull():
                return QIcon(pm)
    except Exception:
        pass
    return widget.style().standardIcon(fallback_sp)

def _make_sidebar_item(text: str, icon: QIcon) -> QListWidgetItem:
    it = QListWidgetItem(icon, text)
    it.setSizeHint(QSize(0, 38))
    return it

# --- Errores / WS URL ---
def _show_err(widget, exc: Exception):
    try:
        if isinstance(exc, requests.RequestException) and exc.response is not None:
            data = exc.response.json()
            msg = data.get("detail") or data
        else:
            msg = str(exc)
    except Exception:
        msg = str(exc)
    QMessageBox.critical(widget, "Error", json.dumps(msg, ensure_ascii=False, indent=2))

def _ws_url_from_api(http_base: str) -> str:
    if http_base.startswith("https://"):
        scheme = "wss://"; base = http_base[len("https://"):]
    elif http_base.startswith("http://"):
        scheme = "ws://";  base = http_base[len("http://"):]
    else:
        scheme = "ws://";  base = http_base
    return f"{scheme}{base}/ws"

# --- Wrappers HTTP locales (se usan en todo el archivo) ---
def api_get(path: str, widget):
    try:
        r = requests.get(f"{API}{path}", timeout=TIMEOUT, headers=_headers())
        r.raise_for_status()
        return r.json()
    except Exception as e:
        _show_err(widget, e)
        return None

def api_post(path: str, payload: dict, widget):
    try:
        r = requests.post(f"{API}{path}", json=payload, timeout=TIMEOUT, headers=_headers())
        r.raise_for_status()
        return r.json()
    except Exception as e:
        _show_err(widget, e)
        return None

def api_delete(path: str, widget, params: dict | None = None):
    try:
        r = requests.delete(f"{API}{path}", params=params or {}, timeout=TIMEOUT, headers=_headers())
        r.raise_for_status()
        return r.json()
    except Exception as e:
        _show_err(widget, e)
        return None

    


# ========= Panel de Administración (Qt) =========





class DlgLogin(QDialog):
    """
    Login:
      - Manual: email_principal + subusuario + password -> POST /auth/login-subuser
      - Google: OAuth -> POST /auth/google/desktop -> elegir subusuario -> POST /auth/login-subuser (modo google)
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Iniciar sesión")
        self.setModal(True)
        self.result_data = None

        self._bg_pm = QPixmap(str(BG_APP_PATH)) if BG_APP_PATH.exists() else None
        self.resize(1100, 650)
        self.setMinimumSize(900, 520)

        root = QHBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(18)

        left  = self._build_left_card()
        right = self._build_right_hero()
        root.addWidget(left, 0)
        root.addWidget(right, 10)

        self._apply_styles()

        self.ed_pwd.returnPressed.connect(self._on_login)
        self.btn_login.clicked.connect(self._on_login)
        self.btn_cancel.clicked.connect(self.reject)
        self.btn_google.clicked.connect(self._on_google_clicked)
        self.btn_register.clicked.connect(self._on_register)

        self._add_toggle_password_action()

    def _apply_styles(self):
        # QSS: Qt NO soporta :root ni overflow.
        self.setStyleSheet("""
        QLabel#LoginSubtitle { color: #6b7280; font-size: 12.5px; }
        QLabel#LoginLead { color: #4b5563; margin-bottom: 6px; }

        QFrame#LoginCard { background: #ffffff; border-radius: 14px; }
        QFrame#HeroWrap  { border-radius: 12px; }

        QLineEdit {
            height: 34px;
            padding: 0 10px;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            background: #fff;
        }
        QLineEdit:focus { border: 1px solid #0b57d0; }

        QPushButton {
            height: 34px;
            padding: 0 14px;
            border-radius: 8px;
            border: 1px solid #e5e7eb;
            background: #f9fafb;
        }
        QPushButton:hover { background: #f3f4f6; }

        QPushButton#GoogleBtn { background: #0b57d0; color: white; border: none; }
        QPushButton#GoogleBtn:hover { background: #0a4ec0; }

        QFrame#Line, QFrame[objectName="Line"] { background: #e5e7eb; height: 1px; }
        """)

    def _add_toggle_password_action(self):
        act = QAction(self)
        act.setIcon(QIcon.fromTheme("view-hidden"))
        act.setToolTip("Mostrar/Ocultar contraseña")

        def toggle():
            self.ed_pwd.setEchoMode(
                QLineEdit.Normal if self.ed_pwd.echoMode() == QLineEdit.Password else QLineEdit.Password
            )
        act.triggered.connect(toggle)
        self.ed_pwd.addAction(act, QLineEdit.TrailingPosition)

    def _build_left_card(self) -> QWidget:
        card = QFrame()
        card.setObjectName("LoginCard")
        card.setFrameShape(QFrame.StyledPanel)

        shadow = QGraphicsDropShadowEffect(blurRadius=24, xOffset=0, yOffset=12)
        shadow.setColor(Qt.black)
        card.setGraphicsEffect(shadow)

        lay = QVBoxLayout(card)
        lay.setContentsMargins(36, 36, 36, 36)
        lay.setSpacing(18)

        header = QHBoxLayout()
        header.setSpacing(12)

        if LOGO_PATH and LOGO_PATH.exists():
            logo_btn = QToolButton()
            logo_btn.setIcon(QIcon(str(LOGO_PATH)))
            logo_btn.setIconSize(QSize(36, 36))
            logo_btn.setCursor(Qt.ArrowCursor)
            logo_btn.setAutoRaise(True)
            header.addWidget(logo_btn, 0, Qt.AlignLeft)
        else:
            logo_lbl = QLabel("•")
            f = QFont(); f.setPointSize(18); f.setBold(True)
            logo_lbl.setFont(f)
            header.addWidget(logo_lbl, 0, Qt.AlignLeft)

        title_wrap = QVBoxLayout()
        lab_title = QLabel("AOP")
        f = QFont(); f.setPointSize(22); f.setBold(True)
        lab_title.setFont(f)

        lab_sub = QLabel("Integramos tus ideas")
        lab_sub.setObjectName("LoginSubtitle")

        title_wrap.addWidget(lab_title)
        title_wrap.addWidget(lab_sub)
        header.addLayout(title_wrap, 1)
        header.addStretch(1)
        lay.addLayout(header)

        lay.addSpacing(6)
        lab_welcome = QLabel("Bienvenido, por favor inicia sesión en tu cuenta.")
        lab_welcome.setObjectName("LoginLead")
        lay.addWidget(lab_welcome)

        form = QFormLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(12)

        self.ed_email_principal = QLineEdit()
        self.ed_email_principal.setPlaceholderText("correo principal (Google)")

        self.ed_subuser = QLineEdit()
        self.ed_subuser.setPlaceholderText("subusuario / etiqueta")

        self.ed_pwd = QLineEdit()
        self.ed_pwd.setPlaceholderText("Contraseña")
        self.ed_pwd.setEchoMode(QLineEdit.Password)

        form.addRow("Email principal", self.ed_email_principal)
        form.addRow("Subusuario", self.ed_subuser)
        form.addRow("Contraseña", self.ed_pwd)
        lay.addLayout(form)

        btns = QHBoxLayout()
        self.btn_login = QPushButton("Entrar")
        self.btn_cancel = QPushButton("Cancelar")
        self.btn_register = QPushButton("Registrarse")
        btns.addWidget(self.btn_login)
        btns.addWidget(self.btn_cancel)
        btns.addWidget(self.btn_register)
        lay.addLayout(btns)

        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setObjectName("Line")
        lay.addWidget(sep)

        self.btn_google = QPushButton("Continuar con Google")
        self.btn_google.setObjectName("GoogleBtn")
        self.btn_google.setEnabled(bool(signin_with_google))
        lay.addWidget(self.btn_google)

        lay.addItem(QSpacerItem(0, 0, QSizePolicy.Minimum, QSizePolicy.Expanding))
        return card

    def _build_right_hero(self) -> QWidget:
        wrap = QFrame()
        wrap.setObjectName("HeroWrap")

        stack = QStackedLayout(wrap)
        stack.setContentsMargins(0, 0, 0, 0)
        stack.setStackingMode(QStackedLayout.StackAll)

        self.hero = QLabel()
        self.hero.setScaledContents(True)
        self.hero.setAlignment(Qt.AlignCenter)
        if BG_LOGIN_PATH.exists():
            self.hero.setPixmap(QPixmap(str(BG_LOGIN_PATH)))
        stack.addWidget(self.hero)

        self.logo_center = QLabel()
        if LOGO_PATH.exists():
            logo_pix = QPixmap(str(LOGO_PATH)).scaled(260, 260, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.logo_center.setPixmap(logo_pix)
        self.logo_center.setAlignment(Qt.AlignCenter)
        stack.addWidget(self.logo_center)

        return wrap

    # ----------------------------
    # Manual login
    # ----------------------------
    def _on_login(self):
        email_principal = (self.ed_email_principal.text() or "").strip()
        usuario = (self.ed_subuser.text() or "").strip()
        pwd = (self.ed_pwd.text() or "").strip()

        if not email_principal or "@" not in email_principal:
            QMessageBox.warning(self, "Login", "Completa un Email principal válido.")
            return
        if not usuario:
            QMessageBox.warning(self, "Login", "Completa el subusuario.")
            return
        if not pwd:
            QMessageBox.warning(self, "Login", "Completa la contraseña.")
            return

        # OJO: si tu backend usa otro path, cámbialo aquí
        r = api_post("/auth/login-subuser", {
            "email_principal": email_principal,
            "usuario": usuario,
            "password": pwd
        }, self)

        if not r or not (r.get("access_token") or r.get("token")):
            QMessageBox.critical(self, "Login", "No fue posible iniciar sesión (token inválido).")
            return

        self.result_data = r
        self.accept()

    # ----------------------------
    # Google login
    # ----------------------------
    def _on_google_clicked(self):
        if not signin_with_google:
            QMessageBox.information(self, "Google", "El inicio con Google no está disponible.")
            return

        try:
            g = _signin_with_google_compat(self) or {}

            # Lo ideal en desktop: tener id_token directo y usar /auth/google
            id_token = (g.get("id_token") or "").strip()

            info = None
            if id_token:
                info = api_post("/auth/google", {"id_token": id_token}, self)
            else:
                # Fallback: si tu helper devuelve code + redirect_uri
                code = (g.get("code") or "").strip()
                redirect_uri = (g.get("redirect_uri") or "").strip()
                if code and redirect_uri:
                    info = api_post("/auth/google/desktop", {"code": code, "redirect_uri": redirect_uri}, self)

            if not info:
                return

            email_principal = (info.get("email_principal") or "").strip()
            subusuarios = info.get("subusuarios") or []

            if not email_principal:
                QMessageBox.warning(self, "Google", "No pude determinar el email principal.")
                return

            if not subusuarios:
                QMessageBox.information(
                    self,
                    "Sin subusuarios",
                    "Tu cuenta principal está verificada, pero no tienes subusuarios creados.\n"
                    "Pídele a un ADMIN que te cree al menos un subusuario (etiqueta) y te asigne rol."
                )
                return

            # elegir subusuario
            usuario_sel = self._pick_subuser(subusuarios)
            if not usuario_sel:
                return

            # pedir contraseña (mantener el login igual que siempre)
            pwd, ok = QInputDialog.getText(
                self,
                "Contraseña del subusuario",
                f"Contraseña para '{usuario_sel}':",
                QLineEdit.Password
            )
            if not ok:
                return
            pwd = (pwd or "").strip()
            if not pwd:
                QMessageBox.warning(self, "Login", "Debes escribir la contraseña.")
                return

            r = api_post("/auth/login-subuser", {
                "email_principal": email_principal,
                "usuario": usuario_sel,
                "password": pwd
            }, self)

            if not r or not (r.get("access_token") or r.get("token")):
                QMessageBox.critical(self, "Login", "No fue posible iniciar sesión (token inválido).")
                return

            self.result_data = r
            self.accept()

        except Exception as ex:
            QMessageBox.critical(self, "Google", f"Fallo al autenticarse con Google.\n{ex}")


    def _pick_subuser(self, subusuarios: list[dict]) -> str | None:
        # subusuarios esperado: [{"usuario":"xxx","rol":"ADMIN","is_active":1}, ...]
        items = []
        for s in subusuarios:
            u = (s.get("usuario") or "").strip()
            rol = (s.get("rol") or "").strip()
            ok = int(s.get("is_active") or 0)
            items.append(f"{u}  ·  {rol or '—'}  ·  {'OK' if ok else 'PENDIENTE'}")

        if not items:
            return None

        chosen, ok = QInputDialog.getItem(self, "Seleccionar subusuario", "Elige un subusuario:", items, 0, False)
        if not ok or not chosen:
            return None

        # recuperar usuario real
        idx = items.index(chosen)
        return (subusuarios[idx].get("usuario") or "").strip() or None

    def _create_subuser_flow(self, email_principal: str) -> str | None:
        usuario, ok = QInputDialog.getText(self, "Crear subusuario", "Nombre de subusuario (etiqueta):")
        if not ok:
            return None
        usuario = (usuario or "").strip()
        if not usuario:
            return None

        password, ok = QInputDialog.getText(self, "Crear subusuario", "Contraseña:", QLineEdit.Password)
        if not ok:
            return None
        password = (password or "").strip()
        if not password:
            return None

        data = api_post("/auth/subuser/register", {
            "email_principal": email_principal,
            "usuario": usuario,
            "nombre": usuario,
            "password": password,
            "confirm_password": password,
            # si tu backend soporta rol solicitado, puedes mandarlo:
            "rol_solicitado": "ADMIN"
        }, self)

        if data and data.get("ok"):
            QMessageBox.information(self, "Subusuario creado", "Creado. Si queda PENDIENTE, debe aprobarlo un ADMIN.")
            return usuario

        QMessageBox.warning(self, "Error", "No se pudo crear el subusuario.")
        return None

    def _on_register(self):
        preset = (self.ed_email_principal.text() or "").strip()
        dlg = DlgRegister(self, preset_email_principal=preset)
        if dlg.exec() == QDialog.Accepted and dlg.result_data:
            self.ed_email_principal.setText(dlg.result_data.get("email_principal", ""))
            self.ed_subuser.setText(dlg.result_data.get("usuario", ""))
            self.ed_pwd.setText(dlg.result_data.get("password", ""))




# =========================
#       DISEÑO  
# =========================

class EventBus(QObject):
    # Señales por tipo de evento + cruda
    orderCreated   = Signal(dict)
    orderFinalized = Signal(dict)
    orderCanceled  = Signal(dict)
    inventoryMoved = Signal(dict)
    rawEvent       = Signal(dict)
    connected      = Signal()
    disconnected   = Signal(str)

class WSClient(QObject):
    def __init__(self, url: str, parent=None):
        super().__init__(parent)
        self.url = url
        self.sock = QWebSocket()
        self.bus = EventBus()
        self._reconnect_ms = 2000

        self.sock.connected.connect(self._on_connected)
        self.sock.disconnected.connect(self._on_disconnected)
        self.sock.textMessageReceived.connect(self._on_text)

        # ping keepalive cada 25s
        self._keep = QTimer(self); self._keep.setInterval(25_000)
        self._keep.timeout.connect(lambda: self.sock.sendTextMessage('{"type":"ping"}'))

        # reconexión
        self._retry = QTimer(self); self._retry.setSingleShot(True)
        self._retry.timeout.connect(self.connect)

    def connect(self):
        try:
            self.sock.open(QUrl(self.url))
        except Exception as e:
            self.bus.disconnected.emit(str(e))
            self._retry.start(self._reconnect_ms)

    def _on_connected(self):
        self._keep.start()
        self.bus.connected.emit()

    def _on_disconnected(self):
        self._keep.stop()
        self.bus.disconnected.emit("closed")
        self._retry.start(self._reconnect_ms)

    def _on_text(self, msg: str):
        try:
            data = json.loads(msg or "{}")
        except Exception:
            return
        self.bus.rawEvent.emit(data)
        typ = (data.get("type") or "").lower()
        payload = data.get("payload") or {}

        if   typ == "order.created":   self.bus.orderCreated.emit(payload)
        elif typ == "order.finalized": self.bus.orderFinalized.emit(payload)
        elif typ == "order.canceled":  self.bus.orderCanceled.emit(payload)
        elif typ.startswith("inventory.") or typ.startswith("inventario."):
            self.bus.inventoryMoved.emit(payload)


# ---------------------- Registro ----------------------
# ---------------------- Registro ----------------------
class DlgRegister(QDialog):
    """
    Crear SUBUSUARIO para un Email principal (Google).

    POST /auth/subuser/register
    payload:
      - email_principal
      - usuario
      - nombre
      - password
      - confirm_password
      - rol_solicitado: ADMIN / DISENO / IMPRESION

    Nota:
      El backend debería dejarlo PENDIENTE / is_active=0 hasta aprobación de Admin.
    """
    def __init__(self, parent=None, preset_email_principal: str | None = None):
        super().__init__(parent)
        self.setWindowTitle("Crear subusuario")
        self.setModal(True)
        self.result_data = None

        form = QFormLayout(self)

        # Campos
        self.ed_email_principal = QLineEdit()
        self.ed_email_principal.setPlaceholderText("correo principal (Google)")
        if preset_email_principal:
            self.ed_email_principal.setText(preset_email_principal)

        self.ed_usuario = QLineEdit()
        self.ed_usuario.setPlaceholderText("subusuario / etiqueta (ej: diseno1)")

        self.ed_nombre = QLineEdit()
        self.ed_nombre.setPlaceholderText("Nombre visible (ej: Juan Pérez)")

        self.ed_pass = QLineEdit()
        self.ed_pass.setEchoMode(QLineEdit.Password)
        self.ed_pass.setPlaceholderText("Contraseña (mínimo 8 caracteres)")

        self.ed_pwd2 = QLineEdit()
        self.ed_pwd2.setEchoMode(QLineEdit.Password)
        self.ed_pwd2.setPlaceholderText("Confirmar contraseña")

        self.cmb_rol = QComboBox()
        self.cmb_rol.addItems(["Impresión", "Diseño", "Administrador"])

        # Botones
        self.btn_create = QPushButton("Crear subusuario")
        self.btn_cancel = QPushButton("Cancelar")

        self.btn_create.clicked.connect(self._create)
        self.btn_cancel.clicked.connect(self.reject)

        # Layout
        form.addRow("Email principal", self.ed_email_principal)
        form.addRow("Subusuario", self.ed_usuario)
        form.addRow("Nombre", self.ed_nombre)
        form.addRow("Contraseña", self.ed_pass)
        form.addRow("Confirmar", self.ed_pwd2)
        form.addRow("Rol solicitado", self.cmb_rol)

        row = QHBoxLayout()
        row.addWidget(self.btn_create)
        row.addWidget(self.btn_cancel)
        form.addRow(row)

        # Enter para crear
        self.ed_pwd2.returnPressed.connect(self._create)

    def _rol_code(self) -> str:
        txt = (self.cmb_rol.currentText() or "").lower()
        if "admin" in txt:
            return "ADMIN"
        if "dise" in txt:   # Diseño / Diseno
            return "DISENO"
        return "IMPRESION"

    def _create(self):
        email_principal = (self.ed_email_principal.text() or "").strip()
        usuario = (self.ed_usuario.text() or "").strip()
        nombre = (self.ed_nombre.text() or "").strip()
        p1 = (self.ed_pass.text() or "").strip()
        p2 = (self.ed_pwd2.text() or "").strip()

        if not email_principal or "@" not in email_principal:
            QMessageBox.warning(self, "Registro", "Escribe un Email principal válido.")
            return
        if not usuario:
            QMessageBox.warning(self, "Registro", "Escribe el subusuario (etiqueta).")
            return
        if not nombre:
            QMessageBox.warning(self, "Registro", "Escribe el nombre.")
            return
        if len(p1) < 8:
            QMessageBox.warning(self, "Registro", "La contraseña debe tener al menos 8 caracteres.")
            return
        if p1 != p2:
            QMessageBox.warning(self, "Registro", "Las contraseñas no coinciden.")
            return

        payload = {
            "email_principal": email_principal,
            "usuario": usuario,
            "nombre": nombre,
            "password": p1,
            "confirm_password": p2,
            "rol_solicitado": self._rol_code(),
        }

        r = api_post("/auth/subuser/register", payload, self)
        if not r:
            return

        # Guardar datos útiles para autollenar el login
        self.result_data = {
            "email_principal": email_principal,
            "usuario": usuario,
            "password": p1,
        }

        QMessageBox.information(
            self,
            "OK",
            "Subusuario creado.\nSi queda PENDIENTE, un ADMIN debe aprobarlo en Administración → Pendientes."
        )
        self.accept()



# ---------------------- Login ----------------------


class CheckableComboBox(QComboBox):
    """QComboBox personalizado con checkboxes para multi-selección."""
    # Señal que se emite cuando cambian los checks
    selectionChanged = Signal()

    def __init__(self, placeholder="Seleccionar...", parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.placeholder = placeholder
        self.lineEdit().setPlaceholderText(placeholder)
        
        self.model = QStandardItemModel(self)
        self.setModel(self.model)
        
        # Evento para cerrar el popup si se hace clic fuera
        self.view().viewport().installEventFilter(self)
        self._changed = False

    def eventFilter(self, widget, event):
        if widget == self.view().viewport():
            if event.type() == QEvent.MouseButtonRelease:
                index = self.view().indexAt(event.pos())
                self.handle_item_pressed(index)
                return True
        return super().eventFilter(widget, event)

    def handle_item_pressed(self, index):
        item = self.model.itemFromIndex(index)
        if item.checkState() == Qt.Checked:
            item.setCheckState(Qt.Unchecked)
        else:
            item.setCheckState(Qt.Checked)
        self._changed = True
        self._update_lineEdit()

    def hidePopup(self):
        # Al cerrar el popup, si hubo cambios, emitimos la señal para filtrar
        super().hidePopup()
        if self._changed:
            self.selectionChanged.emit()
            self._changed = False

    def _update_lineEdit(self):
        texts = []
        for i in range(self.model.rowCount()):
            item = self.model.item(i)
            if item.checkState() == Qt.Checked:
                texts.append(item.text())
        
        if not texts:
            self.lineEdit().setText("")
            self.lineEdit().setPlaceholderText(self.placeholder)
        else:
            self.lineEdit().setText(", ".join(texts))

    def addItem(self, text, data=None):
        item = QStandardItem(text)
        item.setCheckable(True)
        item.setCheckState(Qt.Unchecked)
        if data is not None:
            item.setData(data)
        self.model.appendRow(item)

    def clear(self):
        self.model.clear()
        self.lineEdit().setText("")
        self.lineEdit().setPlaceholderText(self.placeholder)

    def checkedData(self):
        """Devuelve una lista con el 'data' de los items chequeados."""
        data_list = []
        for i in range(self.model.rowCount()):
            item = self.model.item(i)
            if item.checkState() == Qt.Checked:
                data_list.append(item.data())
        return data_list

    def checkedTexts(self):
        """Devuelve una lista con el 'text' de los items chequeados."""
        texts = []
        for i in range(self.model.rowCount()):
            item = self.model.item(i)
            if item.checkState() == Qt.Checked:
                texts.append(item.text())
        return texts

    def checkAll(self, check=True):
        state = Qt.Checked if check else Qt.Unchecked
        for i in range(self.model.rowCount()):
            self.model.item(i).setCheckState(state)
        self._update_lineEdit()




class DropLineEdit(QLineEdit):
    """Campo de texto personalizado que soporta Drag & Drop de archivos o carpetas."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        # Si lo que se arrastra contiene URLs (archivos/carpetas), lo aceptamos
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)

    def dropEvent(self, event):
        # Al soltar, extraemos la ruta y la ponemos en el texto
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            url = event.mimeData().urls()[0]
            ruta = url.toLocalFile() # Convierte la URL a ruta local/red
            if ruta:
                self.setText(ruta)
        else:
            super().dropEvent(event)




class PagDisenio(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()
        self._load_lam_tipos()      # solo tipos; marcas/medidas se cargan al cambiar tipo
        self._on_lam_tipo_changed() # inicializa combos de lam (bloquea si SIN LAMINAR)
        self._load_materials()

    # ---------- UI ----------
    def _pick_ruta(self):
        choice = QMessageBox.question(
            self,
            "Seleccionar ruta",
            "¿Qué deseas adjuntar?\n\nSí = Archivo\nNo = Carpeta",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            QMessageBox.Yes
        )
        if choice == QMessageBox.Cancel:
            return

        # FASE 2: Directorio por defecto (El DISCO D del servidor)
        directorio_base = "D:\\"

        if choice == QMessageBox.Yes:
            # Archivo
            fn, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar archivo de impresión",
                directorio_base,
                "Arte (*.pdf *.ai *.cdr *.eps *.tif *.tiff *.jpg *.jpeg *.png);;Todos (*.*)"
            )
            if fn:
                self.ed_ruta.setText(fn)
        else:
            # Carpeta
            folder = QFileDialog.getExistingDirectory(
                self,
                "Seleccionar carpeta de impresión",
                directorio_base,
                QFileDialog.ShowDirsOnly
            )
            if folder:
                self.ed_ruta.setText(folder)


    def _build_ui(self):
        form = QFormLayout(self)

        # --- RUTA (primera fila) ---
        self.ed_ruta = DropLineEdit()
        self.ed_ruta.setPlaceholderText("Arrastra el archivo aquí o usa 'Buscar...'")
        btn_ruta = QPushButton("Buscar…")
        btn_ruta.clicked.connect(self._pick_ruta)
        fila_ruta = QHBoxLayout()
        fila_ruta.addWidget(self.ed_ruta, 1)
        fila_ruta.addWidget(btn_ruta, 0)
        form.addRow("RUTA", fila_ruta)

        # Material / marca / medidas
        self.cmb_material = QComboBox()
        self.cmb_material.currentIndexChanged.connect(self._on_material_change)
        self.cmb_marca   = QComboBox()
        self.cmb_medida  = QComboBox()
        self.cmb_medida.currentIndexChanged.connect(self._validar_ancho_rollo)
        self.cmb_medida.currentIndexChanged.connect(self._sincronizar_ancho_laminado)

        # Ancho del arte
        self.sb_ancho_arte = QDoubleSpinBox()
        self.sb_ancho_arte.setDecimals(2)
        self.sb_ancho_arte.setMaximum(10000)
        self.sb_ancho_arte.setValue(0.0)
        self.sb_ancho_arte.editingFinished.connect(self._refrescar_sugerencia)

        # Laminado (SIN marca)
        self.cmb_lam_tipo  = QComboBox()
        self.cmb_lam_tipo.currentTextChanged.connect(self._on_lam_tipo_changed)
        self.cmb_lam_ancho = QComboBox()


        # Fecha/otros
        self.dt_entrega = QDateEdit()
        self.dt_entrega.setCalendarPopup(True)
        self.dt_entrega.setDisplayFormat("yyyy-MM-dd")
        self.dt_entrega.setDate(QDate.currentDate())

        self.ed_consec = QLineEdit()
        self.sb_largo  = QDoubleSpinBox(); self.sb_largo.setDecimals(2);  self.sb_largo.setMaximum(1_000_000)
        self.sb_rep    = QSpinBox();        self.sb_rep.setMaximum(10000); self.sb_rep.setValue(0)

        # --- Observaciones (nuevo, DISEÑO) ---
        self.txt_obs = QPlainTextEdit()
        self.txt_obs.setPlaceholderText("Instrucciones para producción (opcional)")
        self.txt_obs.setFixedHeight(64)
        

        form.addRow("Material", self.cmb_material)
        form.addRow("Marca", self.cmb_marca)
        form.addRow("Ancho de la mesa de trabajo (cm)", self.sb_ancho_arte)
        form.addRow("Ancho del rollo (cm)", self.cmb_medida)
        form.addRow("Laminado (tipo)", self.cmb_lam_tipo)
        form.addRow("Laminado (ancho cm)", self.cmb_lam_ancho)
        form.addRow("Fecha de entrega", self.dt_entrega)
        form.addRow("Consecutivo", self.ed_consec)
        form.addRow("Largo (cm)", self.sb_largo)
        form.addRow("Repeticiones", self.sb_rep)
        form.addRow("Observaciones", self.txt_obs)  # <- NUEVO

        btn = QPushButton("Crear orden (PENDIENTE)")
        btn.setProperty("variant", "primary")
        btn.clicked.connect(self._crear)
        form.addRow(btn)
        self.setStyleSheet("""
        QLineEdit, QComboBox, QDateEdit, QDoubleSpinBox, QSpinBox, QPlainTextEdit {
            min-height: 34px; font-size: 13px;
        }
        QDoubleSpinBox::up-button, QSpinBox::up-button, QDateEdit::up-button {
            width: 20px; height: 16px; padding: 0px;
        }
        QDoubleSpinBox::down-button, QSpinBox::down-button, QDateEdit::down-button {
            width: 20px; height: 16px; padding: 0px;
        }
        QComboBox::down-arrow { width: 14px; height: 14px; }
        QToolButton { font-size: 12px; }
        """)

    # ---------- helpers UI ----------
    def _select_combo_value(self, combo: QComboBox, value):
        """Selecciona en combo por 'data' (o por texto si no hay match exacto)."""
        for i in range(combo.count()):
            try:
                if combo.itemData(i) == value or str(combo.itemText(i)) == str(value):
                    combo.setCurrentIndex(i)
                    return
            except Exception:
                pass

    # ---------- carga catálogos ----------
    def _load_lam_tipos(self):
        self.cmb_lam_tipo.clear()
        for t in (api_get("/catalogos/laminados/tipos", self) or []):
            self.cmb_lam_tipo.addItem(t["nombre"], t["id"])

    def _on_lam_tipo_changed(self):
        """Bloqueo y medidas según tipo. Sin marcas."""
        tipo_txt = self.cmb_lam_tipo.currentText().strip()

        # Medidas
        self.cmb_lam_ancho.clear()
        medidas = api_get(f"/catalogos/laminados/medidas?tipo={tipo_txt}", self) or []

        if not medidas:  # SIN LAMINAR
            self.cmb_lam_ancho.setEnabled(False)
            self.cmb_lam_ancho.addItem("0.0", 0.0)
            self.cmb_lam_ancho.setCurrentIndex(0)
        else:
            self.cmb_lam_ancho.setEnabled(True)
            for a in medidas:
                self.cmb_lam_ancho.addItem(str(a["ancho"]), float(a["ancho"]))
                
        # Intentar empatar el ancho del laminado con el rollo actual
        self._sincronizar_ancho_laminado()

    def _sincronizar_ancho_laminado(self):
        """Si el ancho del rollo seleccionado existe en los anchos de laminado, lo asigna."""
        if self.cmb_medida.currentIndex() < 0 or not self.cmb_lam_ancho.isEnabled():
            return
            
        try:
            ancho_rollo = float(self.cmb_medida.currentData())
        except (ValueError, TypeError):
            return
            
        # Busca en el catálogo de laminados si existe el mismo ancho
        for i in range(self.cmb_lam_ancho.count()):
            try:
                ancho_lam = float(self.cmb_lam_ancho.itemData(i))
                if abs(ancho_lam - ancho_rollo) < 0.01:
                    self.cmb_lam_ancho.setCurrentIndex(i)
                    break
            except (ValueError, TypeError):
                pass

    def _load_materials(self):
        data = api_get("/catalogos/materiales", self) or []
        self.cmb_material.clear()
        for r in data:
            self.cmb_material.addItem(r["nombre"], r["id"])
        self._on_material_change()

    def _on_material_change(self):
        if self.cmb_material.currentIndex() < 0:
            return
        mat_id = int(self.cmb_material.currentData())

        # --- Lógica Lona -> Sin Laminar ---
        mat_nombre = self.cmb_material.currentText().strip().upper()
        if "LONA" in mat_nombre or "BANNER" in mat_nombre:
            for i in range(self.cmb_lam_tipo.count()):
                if "SIN LAMINAR" in self.cmb_lam_tipo.itemText(i).upper():
                    self.cmb_lam_tipo.setCurrentIndex(i)
                    break

        # Marcas del material
        self.cmb_marca.clear()
        for m in (api_get(f"/catalogos/marcas?material_id={mat_id}", self) or []):
            self.cmb_marca.addItem(m["nombre"], m["id"])

        # Medidas válidas del rollo para ese material
        self.cmb_medida.clear()
        for a in (api_get(f"/catalogos/medidas?material_id={mat_id}", self) or []):
            self.cmb_medida.addItem(str(a["ancho"]), float(a["ancho"]))

        # intentar sugerir rollo si ya hay un arte escrito
        self._refrescar_sugerencia()

    # ---------- validaciones y sugerencias ----------
    def _refrescar_sugerencia(self):
        """Pide al backend el rollo sugerido con validación estricta (arte < rollo)."""
        if self.cmb_material.currentIndex() < 0 or self.cmb_medida.count() == 0:
            return
        mat_id = int(self.cmb_material.currentData())
        ancho_arte = float(self.sb_ancho_arte.value() or 0)
        if ancho_arte <= 0:
            return
        

        try:
            r = requests.get(
            f"{API}/catalogos/medidas/sugerida",
            params={"material_id": mat_id, "ancho_orden_cm": ancho_arte, "estricto": 1},
            timeout=TIMEOUT,
            headers=_headers(),    
            )
            
            if r.status_code == 200:
                sugerido = float(r.json().get("ancho_rollo_cm"))
                self._select_combo_value(self.cmb_medida, sugerido)
            else:
                # 400: no existe rollo mayor que el arte
                msg = r.json().get("detail", "Medida no válida / material no válido.")
                QMessageBox.warning(self, "Medida no válida", msg)
        except Exception as e:
            _show_err(self, e)

    def _validar_ancho_rollo(self):
        """Si el usuario elige manualmente un rollo < arte, avisamos y proponemos el correcto."""
        if self.cmb_medida.currentIndex() < 0:
            return
        try:
            rollo = float(self.cmb_medida.currentData())
            arte  = float(self.sb_ancho_arte.value() or 0)
            if arte > 0 and rollo <= arte:
                QMessageBox.warning(
                    self,
                    "Medida no válida",
                    "El ancho del arte debe ser estrictamente menor que el ancho del rollo.\n"
                    "Se sugerirá el rollo válido más próximo."
                )
                self._refrescar_sugerencia()
        except Exception:
            pass
    
    def _warn_stock_material(self):
        """
        Consulta existencias del material (material_id, marca_id opcional, ancho_rollo_cm)
        y si stock_cm < (largo_cm * repeticiones) muestra una advertencia.
        No bloquea la creación de la orden.
        """
        try:
            # Selección actual
            mat_id = int(self.cmb_material.currentData())
            mk_id  = int(self.cmb_marca.currentData()) if self.cmb_marca.currentIndex() >= 0 else None
            ancho  = float(self.cmb_medida.currentData())
            largo  = float(self.sb_largo.value())
            rep    = int(self.sb_rep.value())

            requerido_cm = max(0.0, largo) * max(0, rep)

            # Construir query a inventario
            qs = f"/inventario/existencias?material_id={mat_id}&ancho_cm={ancho}"
            if mk_id is not None:
                qs += f"&marca_id={mk_id}"

            rows = api_get(qs, self) or []
            disponible_cm = float(rows[0].get("stock_cm", 0.0)) if rows else 0.0

            if disponible_cm < requerido_cm:
                QMessageBox.warning(
                    self,
                    "Advertencia de stock",
                    (
                        "No hay existencias suficientes del material seleccionado.\n\n"
                        f"Material: {self.cmb_material.currentText()}\n"
                        f"Marca: {self.cmb_marca.currentText() if self.cmb_marca.currentIndex()>=0 else '(sin marca)'}\n"
                        f"Ancho del rollo: {ancho:.1f} cm\n\n"
                        f"Requerido (largo × repeticiones): {requerido_cm:.2f} cm\n"
                        f"Disponible en inventario: {disponible_cm:.2f} cm\n\n"
                        "La orden se creará de todos modos."
                    )
                )
        except Exception:
            # En caso de cualquier problema de red/parsing, no bloqueamos la creación
            pass


    # ---------- crear ----------
    def _crear(self):
        if self.cmb_material.currentIndex() < 0 or self.cmb_medida.currentIndex() < 0:
            QMessageBox.warning(self, "Aviso", "Selecciona material y medida.")
            return
        if not self.ed_consec.text().strip():
            QMessageBox.warning(self, "Aviso", "Ingresa el consecutivo.")
            return
        if self.sb_largo.value() <= 0:
            QMessageBox.warning(self, "Aviso", "El LARGO debe ser mayor a 0.")
            return
            
        # --- FASE 3: Captura y Estandarización de RUTA ---
        ruta_txt = self.ed_ruta.text().strip()
        if not ruta_txt:
            QMessageBox.warning(self, "Aviso", "Debes seleccionar la RUTA del archivo.")
            return

        # Normalizamos las barras (cambia / por \ en Windows)
        ruta_txt = os.path.normpath(ruta_txt)

        # TRADUCCIÓN DE RUTA LOCAL A RUTA DE RED
        # Cambia "NOMBRE_DEL_SERVIDOR" por el nombre real de tu PC servidor en la red o su IP (ej. 192.168.1.50)
        # Cambia "NombreCarpetaCompartida" por el nombre exacto de cómo se llama la carpeta al compartirla.
        # Por ejemplo, si todo el Disco D está compartido como "DISCO D":
        prefijo_red = r"\\NOMBRE_DEL_SERVIDOR\NombreCarpetaCompartida" 
        
        # Si la ruta capturada empieza con D:\, la reemplazamos
        if ruta_txt.upper().startswith(r"D:\\") or ruta_txt.upper().startswith(r"D:"):
            # Quita el 'D:\' y une el resto de la ruta al prefijo de red
            resto_ruta = ruta_txt[3:] if ruta_txt[2] == '\\' else ruta_txt[2:]
            ruta_txt = os.path.join(prefijo_red, resto_ruta)

        # --- FASE 4: Validación de accesibilidad ---
        if not os.path.exists(ruta_txt):
            QMessageBox.warning(
                self, 
                "Ruta inaccesible", 
                f"No se puede acceder a la ruta especificada:\n{ruta_txt}\n\nVerifica la conexión a la red o si el archivo fue movido."
            )
            return

        # Advertencia de existencia (no bloquea)
        self._warn_stock_material()


        payload = {
        "fecha": None,
        "ruta": ruta_txt,
        "material_id": int(self.cmb_material.currentData()),
        "marca_id": int(self.cmb_marca.currentData()) if self.cmb_marca.currentIndex() >= 0 else None,
        "ancho_orden_cm": float(self.sb_ancho_arte.value()) or None,
        "ancho_rollo_cm": float(self.cmb_medida.currentData()),
        "largo": float(self.sb_largo.value()),
        "rep": int(self.sb_rep.value()),
        "consecutivo": self.ed_consec.text().strip(),
        "observaciones": (self.txt_obs.toPlainText().strip() or None),   # <- NUEVO

        # laminado sin marca (enviamos None)
        "lam_tipo_id":  int(self.cmb_lam_tipo.currentData()) if self.cmb_lam_tipo.count() else None,
        "lam_marca_id": None,
        "lam_ancho_cm": float(self.cmb_lam_ancho.currentData()) if self.cmb_lam_ancho.count() else None,

        "fecha_entrega": self.dt_entrega.date().toString("yyyy-MM-dd"),
    }

        r = api_post("/ordenes", payload, self)
        if r:
            QMessageBox.information(self, "OK", f"Orden creada: {json.dumps(r, ensure_ascii=False)}")




# =========================
#     INVENTARIO
# =========================
class _InvBase(QWidget):
    """Base para Materiales / Laminados"""
    def _add_exist_row(self, txt: str):
        self.list_exist.addItem(txt)

    def _clear_exist(self):
        self.list_exist.clear()
        self.list_exist.addItem("No hay existencias para esta combinación.")



class PagInvMaterial(_InvBase):
    def __init__(self):
        super().__init__()
        self._inv_checked = False
        self._inv_ok = True
        self._build()
        self._load_materiales()

    # ====== HTTP directo para inventario (para capturar 404 limpio) ======
    def _api_base(self) -> str:
        # Ajusta aquí si tu variable global se llama distinto.
        base = globals().get("API", "") or ""
        return (base or "").rstrip("/")

    def _api_token(self) -> str:
        # Ajusta aquí si tu token global se llama distinto.
        return globals().get("AUTH_TOKEN", "") or globals().get("API_TOKEN", "") or ""

    def _inv_get(self, path: str, params: dict) -> list[dict]:
        base = self._api_base()
        tok = self._api_token()
        if not base:
            raise RuntimeError("No encuentro API base en global 'API'.")

        url = f"{base}{path}"
        r = requests.get(url, headers=_auth_headers(tok), params=params, timeout=TIMEOUT)

        if r.status_code == 404:
            # router no montado / prefijo incorrecto
            self._inv_ok = False
            return []

        r.raise_for_status()
        data = _safe_json(r)
        if data is None:
            return []
        return data if isinstance(data, list) else (data.get("items") or data.get("rows") or [])

    def _inv_post(self, path: str, payload: dict) -> dict | None:
        base = self._api_base()
        tok = self._api_token()
        url = f"{base}{path}"
        r = requests.post(url, headers=_auth_headers(tok), json=payload, timeout=TIMEOUT)
        if r.status_code == 404:
            self._inv_ok = False
            return None
        r.raise_for_status()
        data = _safe_json(r)
        return data if isinstance(data, dict) else {"ok": True}

    def _inv_delete(self, path: str, params: dict) -> dict | None:
        base = self._api_base()
        tok = self._api_token()
        url = f"{base}{path}"
        r = requests.delete(url, headers=_auth_headers(tok), params=params, timeout=TIMEOUT)
        if r.status_code == 404:
            self._inv_ok = False
            return None
        r.raise_for_status()
        data = _safe_json(r)
        return data if isinstance(data, dict) else {"ok": True}

    def _warn_if_inv_missing(self):
        # muestra SOLO una vez
        if self._inv_checked:
            return
        self._inv_checked = True
        if not self._inv_ok:
            QMessageBox.critical(
                self,
                "Inventario no disponible",
                "El backend respondió 404 a endpoints de inventario.\n\n"
                "Solución:\n"
                "1) Revisa que incluiste el router en FastAPI:\n"
                "   app.include_router(inventario.router, prefix='/inventario')\n"
                "2) Verifica en /docs que existan:\n"
                "   GET /inventario/existencias\n"
                "   POST /inventario/ensure\n"
            )

    # ---------- helpers locales ----------
    def _icon_from_assets_or_std(self, filename: str, fallback_sp: QStyle.StandardPixmap) -> QIcon:
        base = os.path.join(os.path.dirname(__file__), "assets", "icons")
        path = os.path.join(base, filename)
        if os.path.exists(path):
            pm = QPixmap(path)
            if not pm.isNull():
                return QIcon(pm)
        return self.style().standardIcon(fallback_sp)
    

    def _dollar_icon(self) -> QIcon:
        pix = QPixmap(20, 20)
        pix.fill(Qt.transparent)

        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing)

        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        painter.setFont(font)

        painter.drawText(pix.rect(), Qt.AlignCenter, "$")
        painter.end()

        return QIcon(pix)

    



    def _beautify_table(self, tbl: QTableWidget):
        brand = globals().get("COMPANY_YELLOW", "#FFD34D")
        tbl.horizontalHeader().setStyleSheet(
            f"""
            QHeaderView::section {{
                background: {brand};
                color: #222;
                font-weight: 600;
                border: 1px solid #e8c35c;
                padding: 6px 8px;
            }}
            """
        )
        tbl.setAlternatingRowColors(True)
        tbl.setStyleSheet(tbl.styleSheet() + " QTableWidget { gridline-color: #E0E0E0; } ")
        hdr = tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        hdr.setSectionResizeMode(QHeaderView.Interactive)

    # ---------- UI ----------
    def _build(self):
        form = QFormLayout(self)

        self.cmb_mat   = QComboBox(); self.cmb_mat.currentIndexChanged.connect(self._on_mat)
        self.cmb_marca = QComboBox()
        self.cmb_ancho = QComboBox()

        self.chk_all_an = QCheckBox("Mostrar todas las medidas")
        self.chk_all_ma = QCheckBox("Mostrar todas las marcas")

        self.sb_delta = QDoubleSpinBox()
        self.sb_delta.setDecimals(2)
        self.sb_delta.setMinimum(-10_000_000)
        self.sb_delta.setMaximum(10_000_000)

        self.ed_obs   = QLineEdit()

        # ---- costo/cm ----
        self.sb_costo = QDoubleSpinBox()
        self.sb_costo.setDecimals(4)
        self.sb_costo.setValue(0)
        self.sb_costo.setSpecialValueText("0")

        self.sb_costo.setMinimum(0.0)
        self.sb_costo.setMaximum(10_000_000)

        self.btn_costo = QPushButton("Guardar costo/cm")
        self.btn_costo.clicked.connect(self._set_costo)

        self.btn_buscar = QPushButton("Buscar / Mostrar combinaciones"); self.btn_buscar.clicked.connect(self._buscar)
        self.btn_ensure = QPushButton("Crear combinación (0)");          self.btn_ensure.clicked.connect(self._ensure)
        self.btn_mas    = QPushButton("Ajuste +");                       self.btn_mas.clicked.connect(lambda: self._ajuste(+1))
        self.btn_menos  = QPushButton("Ajuste −");                       self.btn_menos.clicked.connect(lambda: self._ajuste(-1))
        self.btn_del    = QPushButton("Eliminar combinación");           self.btn_del.clicked.connect(self._delete)
        self.btn_mov    = QPushButton("Ver últimos movimientos");        self.btn_mov.clicked.connect(self._movs)
        self.btn_ensure.setProperty("variant", "primary")
        self.btn_mas.setProperty("variant", "secondary")
        self.btn_menos.setProperty("variant", "secondary")

        # ---- tabla (5 columnas) ----
        self.tbl_exist = QTableWidget(0, 5)
        self.tbl_exist.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_exist.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tbl_exist.verticalHeader().setVisible(False)
        self.tbl_exist.verticalHeader().setDefaultSectionSize(32)
        self.tbl_exist.itemSelectionChanged.connect(self._sync_costo_from_selected)

        it0 = QTableWidgetItem(self._icon_from_assets_or_std("box.svg",   QStyle.SP_DirIcon),        "Material")
        it1 = QTableWidgetItem(self._icon_from_assets_or_std("tag.svg",   QStyle.SP_FileIcon),       "Marca")
        it2 = QTableWidgetItem(self._icon_from_assets_or_std("ruler.svg", QStyle.SP_ArrowRight),     "Ancho (cm)")
        it3 = QTableWidgetItem(self._icon_from_assets_or_std("chart.svg", QStyle.SP_ComputerIcon),   "Stock (cm)")
        it4 = QTableWidgetItem(self._dollar_icon(), "Costo/cm")

        self.tbl_exist.setHorizontalHeaderItem(0, it0)
        self.tbl_exist.setHorizontalHeaderItem(1, it1)
        self.tbl_exist.setHorizontalHeaderItem(2, it2)
        self.tbl_exist.setHorizontalHeaderItem(3, it3)
        self.tbl_exist.setHorizontalHeaderItem(4, it4)

        self._beautify_table(self.tbl_exist)
        for i, w in enumerate([260, 160, 110, 130, 130]):
            self.tbl_exist.setColumnWidth(i, w)

        # ---- form layout ----
        form.addRow("Material", self.cmb_mat)
        form.addRow("Marca", self.cmb_marca)
        form.addRow("Ancho (cm)", self.cmb_ancho)
        form.addRow(self.chk_all_an)
        form.addRow(self.chk_all_ma)

        form.addRow("Δ Ajuste (cm)", self.sb_delta)
        form.addRow("Costo por cm", self.sb_costo)
        form.addRow(self.btn_costo)

        form.addRow("Observación", self.ed_obs)

        row = QHBoxLayout()
        row.addWidget(self.btn_buscar); row.addWidget(self.btn_ensure)
        row.addWidget(self.btn_mas); row.addWidget(self.btn_menos); row.addWidget(self.btn_del)
        row.addWidget(self.btn_mov); row.addStretch(1)
        form.addRow(row)

        form.addRow(QLabel("Existencias"))
        form.addRow(self.tbl_exist)


    # ---------- carga combos ----------
    def _load_materiales(self):
        self.cmb_mat.clear()
        for r in (api_get("/catalogos/materiales", self) or []):
            self.cmb_mat.addItem(r["nombre"], r["id"])
        self._on_mat()

    def _on_mat(self):
        if self.cmb_mat.currentIndex() < 0:
            return
        mid = int(self.cmb_mat.currentData())

        self.cmb_marca.clear()
        for m in (api_get(f"/catalogos/marcas?material_id={mid}", self) or []):
            self.cmb_marca.addItem(m["nombre"], m["id"])

        self.cmb_ancho.clear()
        for a in (api_get(f"/catalogos/medidas?material_id={mid}", self) or []):
            self.cmb_ancho.addItem(str(a["ancho"]), float(a["ancho"]))

    def _sel(self):
        m_id = int(self.cmb_mat.currentData()) if self.cmb_mat.currentIndex() >= 0 else None
        mk   = int(self.cmb_marca.currentData()) if self.cmb_marca.currentIndex() >= 0 else None
        an   = float(self.cmb_ancho.currentData()) if self.cmb_ancho.currentIndex() >= 0 else None
        return m_id, mk, an


    def _sync_costo_from_selected(self):
        r = self.tbl_exist.currentRow()
        if r < 0:
            return
        it = self.tbl_exist.item(r, 4)
        if not it:
            return
        try:
            self.sb_costo.setValue(float(it.text().replace(",", ".")))
        except Exception:
            pass

    def _set_costo(self):
        mid = int(self.cmb_mat.currentData()) if self.cmb_mat.currentIndex() >= 0 else None
        parsed = self._parse_selected_exist_material()
        if not mid or not parsed:
            QMessageBox.warning(self, "Aviso", "Selecciona una fila en 'Existencias'.")
            return

        mk = parsed["marca_id"]
        an = parsed["ancho_cm"]
        costo = float(self.sb_costo.value())

        payload = {"material_id": int(mid), "ancho_cm": float(an), "costo_cm": float(costo)}
        if mk is not None:
            payload["marca_id"] = int(mk)

        r = self._inv_post("/inventario/costo", payload)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", f"Costo/cm guardado: {costo:.4f}")
            self._buscar()



    # ---------- tabla helpers ----------
    def _add_exist_row(self, material_txt: str, marca_txt: str, ancho_cm: float, stock_cm: float, costo_cm: float):
        r = self.tbl_exist.rowCount()
        self.tbl_exist.insertRow(r)

        def put(c, text, align_right=False):
            it = QTableWidgetItem("" if text is None else str(text))
            if align_right:
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_exist.setItem(r, c, it)

        put(0, material_txt)
        put(1, marca_txt)
        put(2, f"{float(ancho_cm):.1f}", True)
        put(3, f"{float(stock_cm):.2f}", True)
        def _fmt_costo(v):
            v = float(v or 0)
            if v.is_integer():
                return str(int(v))
            return f"{v:.4f}".rstrip("0").rstrip(".")
        put(4, _fmt_costo(costo_cm), True)




    def _clear_exist(self):
        self.tbl_exist.setRowCount(0)

    def _parse_selected_exist_material(self):
        r = self.tbl_exist.currentRow()
        if r < 0:
            return None

        marca_txt = (self.tbl_exist.item(r, 1).text() if self.tbl_exist.item(r, 1) else "").strip()
        ancho_txt = (self.tbl_exist.item(r, 2).text() if self.tbl_exist.item(r, 2) else "").strip()

        try:
            ancho = float(ancho_txt.replace(",", "."))
        except Exception:
            return None

        if not marca_txt or marca_txt.lower().startswith("(sin marca"):
            mk = None
        else:
            mk = None
            for i in range(self.cmb_marca.count()):
                if (self.cmb_marca.itemText(i) or "").strip() == marca_txt:
                    mk = int(self.cmb_marca.itemData(i))
                    break

        return {"marca_id": mk, "ancho_cm": ancho}

    # ---------- acciones ----------
    def _buscar(self):
        mid, mk, an = self._sel()
        if not mid:
            QMessageBox.warning(self, "Aviso", "Selecciona un material.")
            return

        self._clear_exist()

        marcas = [mk]
        anchos = [an]
        if self.chk_all_ma.isChecked():
            marcas = [None] + [self.cmb_marca.itemData(i) for i in range(self.cmb_marca.count())]
        if self.chk_all_an.isChecked():
            anchos = [self.cmb_ancho.itemData(i) for i in range(self.cmb_ancho.count())]

        mat_txt = (self.cmb_mat.currentText() or "").strip()

        def _marca_txt(marca_id):
            if marca_id is None:
                return "(sin marca)"
            for i in range(self.cmb_marca.count()):
                if self.cmb_marca.itemData(i) == marca_id:
                    return self.cmb_marca.itemText(i)
            return f"#{marca_id}"

        added = 0
        for a in anchos:
            if a is None:
                continue
            for m in marcas:
                params = {"material_id": int(mid), "ancho_cm": float(a)}
                if m is not None:
                    params["marca_id"] = int(m)

                rows = self._inv_get("/inventario/existencias", params)
                for rr in rows:
                    self._add_exist_row(
                        mat_txt,
                        _marca_txt(rr.get("marca_id")),
                        float(rr.get("ancho_cm", 0.0)),
                        float(rr.get("stock_cm", 0.0)),
                        float(rr.get("costo_cm", 0.0)),
                    )
                    added += 1

        self._warn_if_inv_missing()
        if added == 0:
            self._clear_exist()


    def _ensure(self):
        mid, mk, an = self._sel()
        if not mid or an is None:
            QMessageBox.warning(self, "Aviso", "Selecciona material y ancho.")
            return

        payload = {
            "material_id": int(mid),
            "marca_id": (int(mk) if self.cmb_marca.currentIndex() >= 0 else None),
            "ancho_cm": float(an),
        }
        r = self._inv_post("/inventario/ensure", payload)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", "Combinación asegurada (stock=0).")
            self._buscar()

    def _ajuste(self, sgn: int):
        mid, mk, an = self._sel()
        if not mid or an is None:
            QMessageBox.warning(self, "Aviso", "Selecciona material y ancho.")
            return

        delta = float(self.sb_delta.value()) * sgn
        if delta == 0:
            QMessageBox.warning(self, "Aviso", "Δ Ajuste debe ser distinto de 0.")
            return

        payload = {
            "material_id": int(mid),
            "marca_id": (int(mk) if self.cmb_marca.currentIndex() >= 0 else None),
            "ancho_cm": float(an),
            "delta_cm": float(delta),
            "obs": (self.ed_obs.text().strip() or None),
            "usuario": None,
        }

        r = self._inv_post("/inventario/ajuste", payload)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", f"Stock actualizado. Nuevo: {r.get('stock_cm','?')}")
            self._buscar()

    def _delete(self):
        mid = int(self.cmb_mat.currentData()) if self.cmb_mat.currentIndex() >= 0 else None
        parsed = self._parse_selected_exist_material()
        if not mid or not parsed:
            QMessageBox.warning(self, "Aviso", "Selecciona una fila en 'Existencias'.")
            return

        mk = parsed["marca_id"]
        an = parsed["ancho_cm"]

        if QMessageBox.question(
            self, "Eliminar combinación",
            f"¿Eliminar la combinación?\nMaterial ID: {mid}\nMarca ID: {mk}\nAncho: {an} cm",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return

        params = {"material_id": int(mid), "ancho_cm": float(an)}
        if mk is not None:
            params["marca_id"] = int(mk)

        r = self._inv_delete("/inventario/delete", params)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", "Combinación eliminada.")
            self._buscar()

    def _movs(self):
        mid, mk, an = self._sel()
        if not mid or an is None:
            QMessageBox.warning(self, "Aviso", "Selecciona material y ancho.")
            return

        params = {"material_id": int(mid), "ancho_cm": float(an), "limit": 20}
        if self.cmb_marca.currentIndex() >= 0:
            params["marca_id"] = int(mk)

        rows = self._inv_get("/inventario/movimientos", params)
        self._warn_if_inv_missing()
        if not rows:
            QMessageBox.information(self, "Movimientos", "No hay movimientos recientes.")
            return

        msg = "\n".join([
            f"{r['fecha']}  {r['tipo']:<9}  {r.get('cantidad_cm',0)} cm | Ord:{r.get('id_orden','-')} [{r.get('observaciones','')}]"
            for r in rows
        ])
        QMessageBox.information(self, "Movimientos (últimos 20)", msg or "(vacío)")



class MoneySpinBox(QDoubleSpinBox):
    def textFromValue(self, value: float) -> str:
        v = float(value or 0.0)
        if v.is_integer():
            return str(int(v))
        return f"{v:.4f}".rstrip("0").rstrip(".")

    def valueFromText(self, text: str) -> float:
        t = (text or "").strip().replace(",", ".")
        try:
            return float(t)
        except Exception:
            return 0.0






class PagInvLaminado(_InvBase):
    def __init__(self):
        super().__init__()
        self._inv_checked = False
        self._inv_ok = True
        self._build()
        self._load_catalogs()

    # ====== HTTP directo para inventario ======
    def _api_base(self) -> str:
        base = globals().get("API", "") or ""
        return (base or "").rstrip("/")

    def _api_token(self) -> str:
        return globals().get("AUTH_TOKEN", "") or globals().get("API_TOKEN", "") or ""

    def _inv_get(self, path: str, params: dict) -> list[dict]:
        base = self._api_base()
        tok = self._api_token()
        if not base:
            raise RuntimeError("No encuentro API base en global 'API'.")

        url = f"{base}{path}"
        r = requests.get(url, headers=_auth_headers(tok), params=params, timeout=TIMEOUT)

        if r.status_code == 404:
            self._inv_ok = False
            return []

        r.raise_for_status()
        data = _safe_json(r)
        if data is None:
            return []
        return data if isinstance(data, list) else (data.get("items") or data.get("rows") or [])

    def _inv_post(self, path: str, payload: dict) -> dict | None:
        base = self._api_base()
        tok = self._api_token()
        url = f"{base}{path}"
        r = requests.post(url, headers=_auth_headers(tok), json=payload, timeout=TIMEOUT)
        if r.status_code == 404:
            self._inv_ok = False
            return None
        r.raise_for_status()
        data = _safe_json(r)
        return data if isinstance(data, dict) else {"ok": True}

    def _inv_delete(self, path: str, params: dict) -> dict | None:
        base = self._api_base()
        tok = self._api_token()
        url = f"{base}{path}"
        r = requests.delete(url, headers=_auth_headers(tok), params=params, timeout=TIMEOUT)
        if r.status_code == 404:
            self._inv_ok = False
            return None
        r.raise_for_status()
        data = _safe_json(r)
        return data if isinstance(data, dict) else {"ok": True}

    def _warn_if_inv_missing(self):
        if self._inv_checked:
            return
        self._inv_checked = True
        if not self._inv_ok:
            QMessageBox.critical(
                self,
                "Inventario no disponible",
                "El backend respondió 404 a endpoints de inventario (laminados).\n\n"
                "Revisa que tu FastAPI incluya el router inventario con prefix '/inventario'\n"
                "y confirma en /docs que exista:\n"
                "GET /inventario/laminados/existencias"
            )

    # ---------- helpers locales ----------
    def _icon_from_assets_or_std(self, filename: str, fallback_sp: QStyle.StandardPixmap) -> QIcon:
        base = os.path.join(os.path.dirname(__file__), "assets", "icons")
        path = os.path.join(base, filename)
        if os.path.exists(path):
            pm = QPixmap(path)
            if not pm.isNull():
                return QIcon(pm)
        return self.style().standardIcon(fallback_sp)
    

    def _dollar_icon(self) -> QIcon:
        pix = QPixmap(20, 20)
        pix.fill(Qt.transparent)

        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing)

        font = QFont()
        font.setBold(True)
        font.setPointSize(11)
        painter.setFont(font)

        painter.drawText(pix.rect(), Qt.AlignCenter, "$")
        painter.end()

        return QIcon(pix)




    def _beautify_table(self, tbl: QTableWidget):
        brand = globals().get("COMPANY_YELLOW", "#FFD34D")
        tbl.horizontalHeader().setStyleSheet(
            f"""
            QHeaderView::section {{
                background: {brand};
                color: #222;
                font-weight: 600;
                border: 1px solid #e8c35c;
                padding: 6px 8px;
            }}
            """
        )
        tbl.setAlternatingRowColors(True)
        tbl.setStyleSheet(tbl.styleSheet() + " QTableWidget { gridline-color: #E0E0E0; } ")
        hdr = tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        hdr.setSectionResizeMode(QHeaderView.Interactive)

    # ---------- UI ----------
    def _build(self):
        form = QFormLayout(self)

        self.cmb_tipo  = QComboBox()
        self.cmb_marca = QComboBox()
        self.cmb_ancho = QComboBox()

        self.chk_all_an = QCheckBox("Mostrar todas las medidas")
        self.chk_all_ma = QCheckBox("Mostrar todas las marcas")

        self.sb_delta = QDoubleSpinBox()
        self.sb_delta.setDecimals(2)
        self.sb_delta.setMinimum(-10_000_000)
        self.sb_delta.setMaximum(10_000_000)

        self.ed_obs   = QLineEdit()

        # ---- costo/cm ----
        self.sb_costo = QDoubleSpinBox()
        self.sb_costo.setDecimals(4)
        self.sb_costo.setValue(0)
        self.sb_costo.setSpecialValueText("0")

        self.sb_costo.setMinimum(0.0)
        self.sb_costo.setMaximum(10_000_000)

        self.btn_costo = QPushButton("Guardar costo/cm")
        self.btn_costo.clicked.connect(self._set_costo)

        self.btn_buscar = QPushButton("Buscar / Mostrar combinaciones"); self.btn_buscar.clicked.connect(self._buscar)
        self.btn_ensure = QPushButton("Crear combinación (0)");          self.btn_ensure.clicked.connect(self._ensure)
        self.btn_mas    = QPushButton("Ajuste +");                       self.btn_mas.clicked.connect(lambda: self._ajuste(+1))
        self.btn_menos  = QPushButton("Ajuste −");                       self.btn_menos.clicked.connect(lambda: self._ajuste(-1))
        self.btn_del    = QPushButton("Eliminar combinación");           self.btn_del.clicked.connect(self._delete)
        self.btn_mov    = QPushButton("Ver últimos movimientos");        self.btn_mov.clicked.connect(self._movs)
        self.btn_ensure.setProperty("variant", "primary")
        self.btn_mas.setProperty("variant", "secondary")
        self.btn_menos.setProperty("variant", "secondary")

        # ---- tabla (5 columnas) ----
        self.tbl_exist = QTableWidget(0, 5)
        self.tbl_exist.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_exist.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tbl_exist.verticalHeader().setVisible(False)
        self.tbl_exist.verticalHeader().setDefaultSectionSize(32)
        self.tbl_exist.itemSelectionChanged.connect(self._sync_costo_from_selected)

        it0 = QTableWidgetItem(self._icon_from_assets_or_std("layers.svg", QStyle.SP_DriveDVDIcon), "Tipo")
        it1 = QTableWidgetItem(self._icon_from_assets_or_std("tag.svg",    QStyle.SP_FileIcon),     "Marca")
        it2 = QTableWidgetItem(self._icon_from_assets_or_std("ruler.svg",  QStyle.SP_ArrowRight),   "Ancho (cm)")
        it3 = QTableWidgetItem(self._icon_from_assets_or_std("chart.svg",  QStyle.SP_ComputerIcon), "Stock (cm)")
        it4 = QTableWidgetItem(self._dollar_icon(), "Costo/cm")

        self.tbl_exist.setHorizontalHeaderItem(0, it0)
        self.tbl_exist.setHorizontalHeaderItem(1, it1)
        self.tbl_exist.setHorizontalHeaderItem(2, it2)
        self.tbl_exist.setHorizontalHeaderItem(3, it3)
        self.tbl_exist.setHorizontalHeaderItem(4, it4)

        self._beautify_table(self.tbl_exist)
        for i, w in enumerate([220, 160, 110, 130, 130]):
            self.tbl_exist.setColumnWidth(i, w)

        # ---- layout ----
        form.addRow("Tipo", self.cmb_tipo)
        form.addRow("Marca", self.cmb_marca)
        form.addRow("Ancho (cm)", self.cmb_ancho)
        form.addRow(self.chk_all_an)
        form.addRow(self.chk_all_ma)

        form.addRow("Δ Ajuste (cm)", self.sb_delta)
        form.addRow("Costo por cm", self.sb_costo)
        form.addRow(self.btn_costo)

        form.addRow("Observación", self.ed_obs)

        row = QHBoxLayout()
        row.addWidget(self.btn_buscar); row.addWidget(self.btn_ensure)
        row.addWidget(self.btn_mas); row.addWidget(self.btn_menos); row.addWidget(self.btn_del)
        row.addWidget(self.btn_mov); row.addStretch(1)
        form.addRow(row)

        form.addRow(QLabel("Existencias Laminado"))
        form.addRow(self.tbl_exist)


    # ---------- combos ----------
    def _load_catalogs(self):
        self.cmb_tipo.clear()
        for t in (api_get("/catalogos/laminados/tipos", self) or []):
            self.cmb_tipo.addItem(t["nombre"], t["id"])

        self.cmb_marca.clear()
        for m in (api_get("/catalogos/laminados/marcas", self) or []):
            self.cmb_marca.addItem(m["nombre"], m["id"])

        self.cmb_ancho.clear()
        for a in (api_get("/catalogos/laminados/medidas", self) or []):
            self.cmb_ancho.addItem(str(a["ancho"]), float(a["ancho"]))

    def _sel(self):
        t = int(self.cmb_tipo.currentData()) if self.cmb_tipo.count() else None
        m = int(self.cmb_marca.currentData()) if self.cmb_marca.count() else None
        a = float(self.cmb_ancho.currentData()) if self.cmb_ancho.count() else None
        return t, m, a


    def _sync_costo_from_selected(self):
        r = self.tbl_exist.currentRow()
        if r < 0:
            return
        it = self.tbl_exist.item(r, 4)
        if not it:
            return
        try:
            self.sb_costo.setValue(float(it.text().replace(",", ".")))
        except Exception:
            pass

    def _set_costo(self):
        t = int(self.cmb_tipo.currentData()) if self.cmb_tipo.currentIndex() >= 0 else None
        parsed = self._parse_selected_exist_lam()
        if not t or not parsed:
            QMessageBox.warning(self, "Aviso", "Selecciona una fila en 'Existencias Laminado'.")
            return

        mk = parsed["lam_marca_id"]
        an = parsed["ancho_cm"]
        costo = float(self.sb_costo.value())

        payload = {"lam_tipo_id": int(t), "ancho_cm": float(an), "costo_cm": float(costo)}
        if mk is not None:
            payload["lam_marca_id"] = int(mk)

        r = self._inv_post("/inventario/laminados/costo", payload)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", f"Costo/cm guardado: {costo:.4f}")
            self._buscar()



    # ---------- tabla helpers ----------
    def _add_exist_row(self, tipo_txt: str, marca_txt: str, ancho_cm: float, stock_cm: float, costo_cm: float):
        r = self.tbl_exist.rowCount()
        self.tbl_exist.insertRow(r)

        def put(c, text, align_right=False):
            it = QTableWidgetItem("" if text is None else str(text))
            if align_right:
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_exist.setItem(r, c, it)

        put(0, tipo_txt)
        put(1, marca_txt)
        put(2, f"{float(ancho_cm):.1f}", True)
        put(3, f"{float(stock_cm):.2f}", True)
        def _fmt_costo(v):
            v = float(v or 0)
            if v.is_integer():
                return str(int(v))
            return f"{v:.4f}".rstrip("0").rstrip(".")
        put(4, _fmt_costo(costo_cm), True)


    def _clear_exist(self):
        self.tbl_exist.setRowCount(0)

    def _parse_selected_exist_lam(self):
        r = self.tbl_exist.currentRow()
        if r < 0:
            return None

        marca_txt = (self.tbl_exist.item(r, 1).text() if self.tbl_exist.item(r, 1) else "").strip()
        ancho_txt = (self.tbl_exist.item(r, 2).text() if self.tbl_exist.item(r, 2) else "").strip()

        try:
            ancho = float(ancho_txt.replace(",", "."))
        except Exception:
            return None

        if not marca_txt or marca_txt.lower().startswith("(sin marca"):
            mk = None
        else:
            mk = None
            for i in range(self.cmb_marca.count()):
                if (self.cmb_marca.itemText(i) or "").strip() == marca_txt:
                    mk = int(self.cmb_marca.itemData(i))
                    break

        return {"lam_marca_id": mk, "ancho_cm": ancho}

    # ---------- acciones ----------
    def _buscar(self):
        t, m, a = self._sel()
        if not t:
            QMessageBox.warning(self, "Aviso", "Selecciona tipo.")
            return

        self._clear_exist()

        marcas = [m]
        anchos = [a]
        if self.chk_all_ma.isChecked():
            marcas = [None] + [self.cmb_marca.itemData(i) for i in range(self.cmb_marca.count())]
        if self.chk_all_an.isChecked():
            anchos = [self.cmb_ancho.itemData(i) for i in range(self.cmb_ancho.count())]

        tipo_map  = {self.cmb_tipo.itemData(i):  self.cmb_tipo.itemText(i)  for i in range(self.cmb_tipo.count())}
        marca_map = {self.cmb_marca.itemData(i): self.cmb_marca.itemText(i) for i in range(self.cmb_marca.count())}
        marca_map[None] = "(sin marca)"

        added = 0
        for an in anchos:
            if an is None:
                continue
            for mk in marcas:
                params = {"lam_tipo_id": int(t), "ancho_cm": float(an)}
                if mk is not None:
                    params["lam_marca_id"] = int(mk)

                rows = self._inv_get("/inventario/laminados/existencias", params)
                for rr in rows:
                    tipo_txt  = tipo_map.get(rr.get("lam_tipo_id"), "LAMINADO")
                    marca_txt = marca_map.get(rr.get("lam_marca_id"), "(sin marca)")

                    self._add_exist_row(
                        tipo_txt,
                        marca_txt,
                        float(rr.get("ancho_cm", 0.0)),
                        float(rr.get("stock_cm", 0.0)),
                        float(rr.get("costo_cm", 0.0)),
                    )
                    added += 1

        self._warn_if_inv_missing()
        if added == 0:
            self._clear_exist()


    def _ensure(self):
        t, m, a = self._sel()
        if not t or a is None:
            QMessageBox.warning(self, "Aviso", "Selecciona tipo y ancho.")
            return

        payload = {
            "lam_tipo_id": int(t),
            "lam_marca_id": (int(m) if self.cmb_marca.currentIndex() >= 0 else None),
            "ancho_cm": float(a),
        }

        r = self._inv_post("/inventario/laminados/ensure", payload)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", "Combinación asegurada (stock=0).")
            self._buscar()

    def _ajuste(self, sgn: int):
        t, m, a = self._sel()
        if not t or a is None:
            QMessageBox.warning(self, "Aviso", "Selecciona tipo y ancho.")
            return

        delta = float(self.sb_delta.value()) * sgn
        if delta == 0:
            QMessageBox.warning(self, "Aviso", "Δ Ajuste debe ser distinto de 0.")
            return

        payload = {
            "lam_tipo_id": int(t),
            "lam_marca_id": (int(m) if self.cmb_marca.currentIndex() >= 0 else None),
            "ancho_cm": float(a),
            "delta_cm": float(delta),
            "obs": (self.ed_obs.text().strip() or None),
            "usuario": None,
        }

        r = self._inv_post("/inventario/laminados/ajuste", payload)
        self._warn_if_inv_missing()

        if r:
            # Si backend devuelve stock_cm, lo mostramos
            nuevo = r.get("stock_cm", None)
            if nuevo is None:
                QMessageBox.information(self, "OK", "Stock actualizado.")
            else:
                QMessageBox.information(self, "OK", f"Stock actualizado. Nuevo: {nuevo}")
            self._buscar()


    def _delete(self):
        t = int(self.cmb_tipo.currentData()) if self.cmb_tipo.currentIndex() >= 0 else None
        parsed = self._parse_selected_exist_lam()
        if not t or not parsed:
            QMessageBox.warning(self, "Aviso", "Selecciona una fila en 'Existencias Laminado'.")
            return

        m = parsed["lam_marca_id"]
        a = parsed["ancho_cm"]

        if QMessageBox.question(
            self, "Eliminar combinación",
            f"¿Eliminar la combinación?\nTipo ID: {t}\nMarca ID: {m}\nAncho: {a} cm",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return

        params = {"lam_tipo_id": int(t), "ancho_cm": float(a)}
        if m is not None:
            params["lam_marca_id"] = int(m)

        r = self._inv_delete("/inventario/laminados/delete", params)
        self._warn_if_inv_missing()
        if r:
            QMessageBox.information(self, "OK", "Combinación eliminada.")
            self._buscar()

    def _movs(self):
        t, m, a = self._sel()
        if not t or a is None:
            QMessageBox.warning(self, "Aviso", "Selecciona tipo y ancho.")
            return

        params = {"lam_tipo_id": int(t), "ancho_cm": float(a), "limit": 20}
        if self.cmb_marca.currentIndex() >= 0:
            params["lam_marca_id"] = int(m)

        rows = self._inv_get("/inventario/laminados/movimientos", params)
        self._warn_if_inv_missing()
        if not rows:
            QMessageBox.information(self, "Movimientos", "No hay movimientos recientes.")
            return

        msg = "\n".join([
            f"{r['fecha']}  {r['tipo']:<9}  {r.get('cantidad_cm',0)} cm | Ord:{r.get('id_orden','-')} [{r.get('observaciones','')}]"
            for r in rows
        ])
        QMessageBox.information(self, "Movimientos (últimos 20)", msg or "(vacío)")

class PagAdmin(QWidget):
    def __init__(self, base_url: str, token: str, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)

        tabs = QTabWidget()
        tabs.addTab(
            AdminSolicitudesPendientesPage(base_url, token, self),
            "Pendientes"
        )
        tabs.addTab(
            AdminSolicitudesFrame(self, base_url, token),
            "Usuarios"
        )

        lay.addWidget(tabs)


class PagAltasCatalogo(QWidget):
    catalog_changed = Signal(str)

    def __init__(self):
        super().__init__()
        self._build()
        self._connect_signals()
        self._load_data()

    def _build(self):
        main_lay = QVBoxLayout(self)

        # --- SECCIÓN 1: SELECTOR DE CATÁLOGO ---
        group_sel = QGroupBox("1. Seleccionar Catálogo a Gestionar")
        lay_sel = QHBoxLayout(group_sel)
        lay_sel.setContentsMargins(12, 16, 12, 16)

        self.cmb_tipo = QComboBox()
        self.cmb_tipo.addItems(["Material", "Laminado"])
        self.cmb_tipo.setMinimumWidth(150)

        self.lbl_subcat = QLabel("Subcategoría:")
        self.cmb_subcat = QComboBox()
        self.cmb_subcat.addItems(["Adhesivo", "Lona"])
        self.cmb_subcat.setMinimumWidth(150)

        lay_sel.addWidget(QLabel("Catálogo principal:"))
        lay_sel.addWidget(self.cmb_tipo)
        lay_sel.addSpacing(30)
        lay_sel.addWidget(self.lbl_subcat)
        lay_sel.addWidget(self.cmb_subcat)
        lay_sel.addStretch()

        # --- SECCIÓN 2: COLUMNAS DE GESTIÓN (Interfaz Limpia) ---
        lay_cols = QHBoxLayout()
        lay_cols.setSpacing(15)

        # Estilo para unificar las listas
        css_lists = """
            QListWidget { border: 1px solid #d0d0d0; border-radius: 6px; background: #ffffff; }
            QListWidget::item { padding: 8px; border-bottom: 1px solid #f0f0f0; color: #333; }
            QListWidget::item:selected { background: #e3f2fd; color: #00695c; font-weight: bold; border-radius: 4px;}
        """

        # --- COLUMNA 1: ITEMS (Nombres) ---
        group_items = QGroupBox("Items Registrados")
        lay_i = QVBoxLayout(group_items)
        self.list_items = QListWidget()
        self.list_items.setStyleSheet(css_lists)
        lay_i.addWidget(self.list_items)
        
        row_i = QHBoxLayout()
        self.ed_item = QLineEdit()
        self.ed_item.setPlaceholderText("Nuevo item...")
        self.btn_add_item = QPushButton("➕ Agregar")
        self.btn_add_item.setProperty("variant", "primary")
        row_i.addWidget(self.ed_item)
        row_i.addWidget(self.btn_add_item)
        lay_i.addLayout(row_i)
        
        self.chk_force_item = QCheckBox("Forzar eliminación (Borra stock físico)")
        lay_i.addWidget(self.chk_force_item)
        self.btn_del_item = QPushButton("🗑️ Eliminar Item Seleccionado")
        self.btn_del_item.setStyleSheet("color: #c62828; font-weight: bold;")
        lay_i.addWidget(self.btn_del_item)

        # --- COLUMNA 2: MARCAS ---
        group_marcas = QGroupBox("Marcas del Catálogo")
        lay_m = QVBoxLayout(group_marcas)
        self.list_marcas = QListWidget()
        self.list_marcas.setStyleSheet(css_lists)
        lay_m.addWidget(self.list_marcas)
        
        row_m = QHBoxLayout()
        self.ed_marca = QLineEdit()
        self.ed_marca.setPlaceholderText("Nueva marca...")
        self.btn_add_marca = QPushButton("➕ Agregar")
        row_m.addWidget(self.ed_marca)
        row_m.addWidget(self.btn_add_marca)
        lay_m.addLayout(row_m)
        
        self.btn_del_marca = QPushButton("🗑️ Eliminar Marca Seleccionada")
        self.btn_del_marca.setStyleSheet("color: #c62828;")
        lay_m.addWidget(self.btn_del_marca)

        # --- COLUMNA 3: MEDIDAS ---
        group_medidas = QGroupBox("Medidas Base (Ancho rollos)")
        lay_w = QVBoxLayout(group_medidas)
        self.list_medidas = QListWidget()
        self.list_medidas.setStyleSheet(css_lists)
        lay_w.addWidget(self.list_medidas)
        
        row_w = QHBoxLayout()
        self.sb_medida = QDoubleSpinBox()
        self.sb_medida.setDecimals(2)
        self.sb_medida.setMaximum(10000)
        self.sb_medida.setSuffix(" cm")
        self.sb_medida.setButtonSymbols(QDoubleSpinBox.NoButtons) # Estética limpia
        self.btn_add_medida = QPushButton("➕ Agregar")
        row_w.addWidget(self.sb_medida)
        row_w.addWidget(self.btn_add_medida)
        lay_w.addLayout(row_w)
        
        self.btn_del_medida = QPushButton("🗑️ Eliminar Medida Seleccionada")
        self.btn_del_medida.setStyleSheet("color: #c62828;")
        lay_w.addWidget(self.btn_del_medida)

        # Ensamblaje
        lay_cols.addWidget(group_items)
        lay_cols.addWidget(group_marcas)
        lay_cols.addWidget(group_medidas)

        main_lay.addWidget(group_sel, 0)
        main_lay.addLayout(lay_cols, 1)


    def _connect_signals(self):
        # Cuando cambia el catálogo principal
        self.cmb_tipo.currentTextChanged.connect(self._on_tipo_changed)
        self.cmb_subcat.currentTextChanged.connect(self._load_data)

        # Botones de agregar
        self.btn_add_item.clicked.connect(self._add_item)
        self.ed_item.returnPressed.connect(self._add_item)
        self.btn_add_marca.clicked.connect(self._add_marca)
        self.ed_marca.returnPressed.connect(self._add_marca)
        self.btn_add_medida.clicked.connect(self._add_medida)

        # Botones de eliminar
        self.btn_del_item.clicked.connect(self._del_item)
        self.btn_del_marca.clicked.connect(self._del_marca)
        self.btn_del_medida.clicked.connect(self._del_medida)

    def _es_adhesivo(self, nombre):
        n = (nombre or "").upper()
        return any(tok in n for tok in ("ADH", "ADHESIVO", "VINIL", "VINILO", "STICKER"))

    def _on_tipo_changed(self, txt: str):
        es_mat = (txt == "Material")
        self.lbl_subcat.setVisible(es_mat)
        self.cmb_subcat.setVisible(es_mat)
        self._load_data()

    # --- LÓGICA DE CARGA (Inteligente y robusta) ---
    # --- LÓGICA DE CARGA (Inteligente y robusta) ---
    def _load_data(self):
        self.list_items.clear()
        self.list_marcas.clear()
        self.list_medidas.clear()

        tipo = self.cmb_tipo.currentText()
        subcat = self.cmb_subcat.currentText()

        # 1. CARGAR ITEMS
        if tipo == "Material":
            mats = api_get("/catalogos/materiales", self) or []
            for m in mats:
                is_adh = self._es_adhesivo(m["nombre"])
                if (subcat == "Adhesivo" and is_adh) or (subcat == "Lona" and not is_adh):
                    it = QListWidgetItem(m["nombre"])
                    it.setData(Qt.UserRole, m["id"])
                    self.list_items.addItem(it)
        else:
            lams = api_get("/catalogos/laminados/tipos", self) or []
            for l in lams:
                it = QListWidgetItem(l["nombre"])
                it.setData(Qt.UserRole, l["id"])
                self.list_items.addItem(it)

        # 2. CARGAR MARCAS
        if tipo == "Material":
            seen_m = set()
            marcas = []
            all_mats = api_get("/catalogos/materiales", self) or []
            for m in all_mats:
                b = api_get(f"/catalogos/marcas?material_id={m['id']}", self) or []
                for x in b:
                    if x["id"] not in seen_m:
                        seen_m.add(x["id"])
                        marcas.append(x)
        else:
            marcas = api_get("/catalogos/laminados/marcas", self) or []

        for m in marcas:
            it = QListWidgetItem(m["nombre"])
            it.setData(Qt.UserRole, m["id"])
            self.list_marcas.addItem(it)

        # 3. CARGAR MEDIDAS (Corregido para evitar el error 422 del backend)
        if tipo == "Material":
            seen_w = set()
            medidas = []
            all_mats = api_get("/catalogos/materiales", self) or []
            for m in all_mats:
                is_adh = self._es_adhesivo(m["nombre"])
                # Solo pedimos las medidas si el material pertenece a la subcategoría actual
                if (subcat == "Adhesivo" and is_adh) or (subcat == "Lona" and not is_adh):
                    w_arr = api_get(f"/catalogos/medidas?material_id={m['id']}", self) or []
                    for x in w_arr:
                        try:
                            w_val = float(x["ancho"])
                            if w_val not in seen_w:
                                seen_w.add(w_val)
                                medidas.append({"ancho": w_val})
                        except Exception:
                            pass
        else:
            medidas = api_get("/catalogos/laminados/medidas", self) or []

        medidas.sort(key=lambda x: float(x["ancho"]))
        for m in medidas:
            w_val = float(m["ancho"])
            it = QListWidgetItem(f"{w_val:.2f} cm")
            it.setData(Qt.UserRole, w_val)
            self.list_medidas.addItem(it)


    # --- LÓGICA DE CREACIÓN ---
    def _add_item(self):
        nombre = self.ed_item.text().strip().upper()
        if not nombre: return
        tipo = self.cmb_tipo.currentText()
        
        if tipo == "Material":
            if self.cmb_subcat.currentText() == "Lona" and not nombre.startswith("LONA"):
                nombre = f"LONA {nombre}"
            r = api_post("/catalogos/materiales/create", {"nombre": nombre}, self)
        else:
            r = api_post("/catalogos/laminados/tipos/create", {"nombre": nombre}, self)

        if r:
            self.ed_item.clear()
            self._load_data()
            self.catalog_changed.emit(tipo)

    def _add_marca(self):
        nombre = self.ed_marca.text().strip().upper()
        if not nombre: return
        tipo = self.cmb_tipo.currentText()
        
        if tipo == "Material":
            r = api_post("/catalogos/marcas/create", {"nombre": nombre}, self)
        else:
            r = api_post("/catalogos/laminados/marcas/create", {"nombre": nombre}, self)
            
        if r:
            self.ed_marca.clear()
            self._load_data()
            self.catalog_changed.emit(tipo)

    def _add_medida(self):
        val = self.sb_medida.value()
        if val <= 0: return
        tipo = self.cmb_tipo.currentText()
        
        if tipo == "Material":
            cat = "LONA" if self.cmb_subcat.currentText() == "Lona" else "ADHESIVO"
            r = api_post("/catalogos/medidas/add", {"categoria": cat, "ancho_cm": val}, self)
        else:
            r = api_post("/catalogos/medidas/add", {"categoria": "ADHESIVO", "ancho_cm": val}, self)
            
        if r:
            self.sb_medida.setValue(0)
            self._load_data()
            self.catalog_changed.emit(tipo)


    # --- LÓGICA DE ELIMINACIÓN ---
    def _del_item(self):
        it = self.list_items.currentItem()
        if not it:
            QMessageBox.warning(self, "Aviso", "Selecciona un Item de la lista para eliminarlo.")
            return
            
        item_id = it.data(Qt.UserRole)
        force = 1 if self.chk_force_item.isChecked() else 0
        tipo = self.cmb_tipo.currentText()

        if QMessageBox.question(self, "Eliminar", f"¿Estás seguro de eliminar '{it.text()}'?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            if tipo == "Material":
                r = api_delete(f"/catalogos/materiales/{item_id}", self, params={"force": force})
            else:
                r = api_delete(f"/catalogos/laminados/tipos/{item_id}", self)
            if r:
                self._load_data()
                self.catalog_changed.emit(tipo)

    def _del_marca(self):
        it = self.list_marcas.currentItem()
        if not it:
            QMessageBox.warning(self, "Aviso", "Selecciona una Marca de la lista.")
            return
            
        m_id = it.data(Qt.UserRole)
        tipo = self.cmb_tipo.currentText()
        
        if QMessageBox.question(self, "Eliminar", f"¿Eliminar la marca '{it.text()}' del sistema?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            if tipo == "Material":
                r = api_delete(f"/catalogos/marcas/{m_id}", self)
            else:
                r = api_delete(f"/catalogos/laminados/marcas/{m_id}", self)
            if r:
                self._load_data()
                self.catalog_changed.emit(tipo)

    def _del_medida(self):
        it = self.list_medidas.currentItem()
        if not it:
            QMessageBox.warning(self, "Aviso", "Selecciona una Medida de la lista.")
            return
            
        w_val = it.data(Qt.UserRole)
        tipo = self.cmb_tipo.currentText()
        cat = "LONA" if (tipo == "Material" and self.cmb_subcat.currentText() == "Lona") else "ADHESIVO"

        if QMessageBox.question(self, "Eliminar", f"¿Eliminar el ancho de {w_val} cm?", QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            r = api_delete("/catalogos/medidas", self, params={"categoria": cat, "ancho_cm": w_val})
            if r:
                self._load_data()
                self.catalog_changed.emit(tipo)
############################################3

from PySide6.QtWidgets import QWidget, QTabWidget, QVBoxLayout

class PagInventario(QWidget):
    def __init__(self):
        super().__init__()

        self.tabs = QTabWidget(self)

        # ✅ Guardamos referencias
        self.pag_inv_mat = PagInvMaterial()
        self.pag_inv_lam = PagInvLaminado()
        self.pag_altas   = PagAltasCatalogo()

        self.tabs.addTab(self.pag_inv_mat, "Materiales")
        self.tabs.addTab(self.pag_inv_lam, "Laminados")
        self.tabs.addTab(self.pag_altas,   "Altas")

        lay = QVBoxLayout(self)
        lay.addWidget(self.tabs)

        # ✅ Conectar señal de Altas -> refrescar inventario
        self.pag_altas.catalog_changed.connect(self._on_catalog_changed)

        # (opcional) refrescar al cambiar de pestaña
        # self.tabs.currentChanged.connect(self._on_tab_changed)

    def _on_catalog_changed(self, tipo: str):
        tipo = (tipo or "").strip().lower()
        if tipo == "material":
            self.pag_inv_mat._load_materiales()
            # opcional: refrescar tabla visible
            # self.pag_inv_mat._buscar()
        elif tipo == "laminado":
            self.pag_inv_lam._load_catalogs()
            # opcional:
            # self.pag_inv_lam._buscar()
        else:
            # fallback
            self.pag_inv_mat._load_materiales()
            self.pag_inv_lam._load_catalogs()

    # opcional
    def _on_tab_changed(self, _idx: int):
        # refresca catálogos al entrar a pestañas
        self.pag_inv_mat._load_materiales()
        self.pag_inv_lam._load_catalogs()



class PagReportes(QWidget):

    HEADERS = [
        "ID", "Estado", "Consumo esp.(cm)", "Fecha", "Entrega", "Consec",
        "Material", "Ancho roll", "Largo", "Rep",
        "Impreso(cm)", "Desp.L", "Desp.A",
        "Lam.tipo", "Lam.marca", "Lam.ancho", "Lam.consumo",
        "Merma cancel(cm)",
        "Costo Mat", "Costo Lam", "Costo Total",  # ✅ NUEVAS
        "Ruta", "Abrir", "Guardar"
    ]




        # --- columnas de costos ---
    COL_COSTO_MAT   = HEADERS.index("Costo Mat")
    COL_COSTO_LAM   = HEADERS.index("Costo Lam")
    COL_COSTO_TOTAL = HEADERS.index("Costo Total")
    COST_COLS = (COL_COSTO_MAT, COL_COSTO_LAM, COL_COSTO_TOTAL)

    COST_BG_HEX = "#E6D14A"
    COST_FG_HEX = "#443F3F"


    def _fmt_cop(self, x: float) -> str:
        """
        Formato pesos colombianos:
        - miles con punto: 1.234.567
        - decimales con coma (máx 2): 1.234,5 / 1.234,56
        - si no hay decimales significativos: no los muestra
        - incluye signo $
        """
        try:
            v = float(x or 0.0)
        except Exception:
            v = 0.0

        # Evitar "-0"
        if abs(v) < 1e-9:
            v = 0.0

        sign = "-" if v < 0 else ""
        v = abs(v)

        v = round(v, 2)
        is_int = abs(v - int(v)) < 1e-9

        if is_int:
            s = f"{int(v):,}".replace(",", ".")  # 1,234 -> 1.234
            return f"{sign}${s}"

        # 2 decimales máximo, pero sin ruido (quita ceros finales)
        s = f"{v:,.2f}".rstrip("0").rstrip(".")  # "1,234.50" -> "1,234.5"
        # Swap a formato COL: miles '.' y decimales ','
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{sign}${s}"


    def _set_cost_cell(self, row: int, col: int, value: float):
        from PySide6.QtCore import Qt
        from PySide6.QtGui import QColor
        from PySide6.QtWidgets import QTableWidgetItem

        try:
            v = float(value or 0.0)
        except Exception:
            v = 0.0

        it = QTableWidgetItem(self._fmt_cop(v))
        it.setData(Qt.UserRole, v)

        # alineación derecha
        it.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)

        # estilo suave, sin negrilla
        it.setBackground(QColor(self.COST_BG_HEX))
        it.setForeground(QColor(self.COST_FG_HEX))

        # no editable
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)

        self.tbl.setItem(row, col, it)


    def _style_cost_headers(self):
        """Pinta los headers de las 3 columnas de costos."""
        from PySide6.QtGui import QColor

        for c in self.COST_COLS:
            h = self.tbl.horizontalHeaderItem(c)
            if not h:
                continue
            h.setBackground(QColor(self.COST_BG_HEX))
            h.setForeground(QColor(self.COST_FG_HEX))
            f = h.font()
            f.setBold(True)
            h.setFont(f)







    def _pick_cost(self, row: dict) -> float:
        # Soporta distintos nombres que pueda devolver tu API
        for k in ("costo_cm", "costo", "costo_por_cm", "costo_cm2", "costo_unitario"):
            v = row.get(k, None)
            if v is None:
                continue
            try:
                return float(v)
            except Exception:
                pass
        return 0.0


    def _get_costo_cm_material_ids(self, material_id: int | None, marca_id: int | None, ancho_cm: float) -> float:
        if not material_id:
            return 0.0

        key = (int(material_id), int(marca_id) if marca_id is not None else None, round(float(ancho_cm), 2))
        if key in self._cost_cache_mat:
            return float(self._cost_cache_mat[key] or 0.0)

        params = {"material_id": int(material_id), "ancho_cm": float(ancho_cm)}
        if marca_id is not None:
            params["marca_id"] = int(marca_id)

        url = "/inventario/existencias?" + urlencode(params)
        rows = api_get(url, self) or []
        costo = float(rows[0].get("costo_cm") or 0.0) if rows else 0.0

        self._cost_cache_mat[key] = costo
        return costo

    def _abrir_ruta(self, ruta: str | None):
        if not ruta:
            QMessageBox.information(self, "Abrir", "La orden no tiene RUTA asociada.")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(ruta))

    def __init__(self):
        super().__init__()
        self._build()
        self._load()

    def _col_settings_key(self) -> str:
        return "reportes/columns_visible"

    def _default_visible_cols(self) -> list[int]:
        return list(range(len(self.HEADERS)))

    def _load_visible_cols(self) -> list[int]:
        s = QSettings()
        raw = s.value(self._col_settings_key(), None)
        if not raw:
            return self._default_visible_cols()
        try:
            parts = [p.strip() for p in str(raw).split(",") if p.strip() != ""]
            cols = [int(p) for p in parts]
            cols = [c for c in cols if 0 <= c < len(self.HEADERS)]
            return cols if cols else self._default_visible_cols()
        except Exception:
            return self._default_visible_cols()

    def _save_visible_cols(self):
        s = QSettings()
        visible = [i for i in range(self.tbl.columnCount()) if not self.tbl.isColumnHidden(i)]
        s.setValue(self._col_settings_key(), ",".join(str(i) for i in visible))

    def _apply_columns_visibility(self, visible_cols: list[int]):
        visible_set = set(visible_cols or [])
        for i in range(self.tbl.columnCount()):
            self.tbl.setColumnHidden(i, i not in visible_set)

    def _build_columns_menu(self):
        self.btn_cols = QToolButton()
        self.btn_cols.setText("Columnas ▾")
        self.btn_cols.setPopupMode(QToolButton.InstantPopup)

        menu = QMenu(self.btn_cols)
        self._col_actions = []

        act_all = menu.addAction("Mostrar todas")
        act_none = menu.addAction("Ocultar todas")
        menu.addSeparator()

        def show_all():
            for a in self._col_actions:
                a.setChecked(True)
            self._on_columns_menu_changed()

        def hide_all():
            for a in self._col_actions:
                a.setChecked(False)
            self._on_columns_menu_changed()

        act_all.triggered.connect(show_all)
        act_none.triggered.connect(hide_all)

        for idx, title in enumerate(self.HEADERS):
            a = menu.addAction(title)
            a.setCheckable(True)
            a.setChecked(True)
            a.toggled.connect(lambda _checked, i=idx: self._on_column_toggled(i))
            self._col_actions.append(a)

        self.btn_cols.setMenu(menu)

        visible_cols = self._load_visible_cols()
        vis = set(visible_cols)
        for i, a in enumerate(self._col_actions):
            a.blockSignals(True)
            a.setChecked(i in vis)
            a.blockSignals(False)

        self._on_columns_menu_changed()

    def _on_column_toggled(self, col_index: int):
        checked = self._col_actions[col_index].isChecked()
        self.tbl.setColumnHidden(col_index, not checked)
        self._save_visible_cols()

    def _on_columns_menu_changed(self):
        if not hasattr(self, "tbl") or self.tbl is None:
            return
        for i, a in enumerate(self._col_actions):
            self.tbl.setColumnHidden(i, not a.isChecked())
        self._save_visible_cols()


    def _norm(self, s: str | None) -> str:
        return " ".join((s or "").strip().upper().split())

    def _fmt_money(self, x: float) -> str:
        try:
            v = float(x or 0.0)
        except Exception:
            v = 0.0
        if abs(v - int(v)) < 1e-9:
            return str(int(v))
        return f"{v:.2f}".rstrip("0").rstrip(".")

    def _safe_float(self, v) -> float:
        try:
            return float(v)
        except Exception:
            return 0.0

    def _init_cost_maps_if_needed(self):
        if not self._mat_name_to_id:
            mats = api_get("/catalogos/materiales", self) or []
            self._mat_name_to_id = {self._norm(m.get("nombre")): int(m.get("id")) for m in mats if m.get("id")}

        if not self._lam_tipo_name_to_id:
            tipos = api_get("/catalogos/laminados/tipos", self) or []
            self._lam_tipo_name_to_id = {self._norm(t.get("nombre")): int(t.get("id")) for t in tipos if t.get("id")}

        if not self._lam_marca_name_to_id:
            marcas = api_get("/catalogos/laminados/marcas", self) or []
            self._lam_marca_name_to_id = {self._norm(m.get("nombre")): int(m.get("id")) for m in marcas if m.get("id")}

    def _material_brand_id(self, mat_id: int, marca_nombre: str | None) -> int | None:
        nombre = self._norm(marca_nombre)
        if not nombre:
            return None

        if mat_id not in self._mat_brand_by_matid:
            rows = api_get(f"/catalogos/marcas?material_id={int(mat_id)}", self) or []
            self._mat_brand_by_matid[mat_id] = {self._norm(r.get("nombre")): int(r.get("id")) for r in rows if r.get("id")}

        return self._mat_brand_by_matid[mat_id].get(nombre)

    from urllib.parse import urlencode

    def _get_costo_cm_material(self, material_nombre: str | None, marca_nombre: str | None, ancho_cm: float) -> float:
        self._init_cost_maps_if_needed()

        # ✅ si viene ID en vez de nombre
        mid = None
        try:
            if material_nombre is not None and str(material_nombre).strip().isdigit():
                mid = int(material_nombre)
        except Exception:
            mid = None
        if mid is None:
            mid = self._mat_name_to_id.get(self._norm(material_nombre))
        if not mid:
            return 0.0

        marca_id = self._material_brand_id(mid, marca_nombre)  # puede ser None
        key = (int(mid), int(marca_id) if marca_id is not None else None, round(float(ancho_cm), 2))
        if key in self._cost_cache_mat:
            return float(self._cost_cache_mat[key] or 0.0)

        params = {"material_id": int(mid), "ancho_cm": float(ancho_cm)}
        if marca_id is not None:
            params["marca_id"] = int(marca_id)

        url = "/inventario/existencias?" + urlencode(params)
        rows = api_get(url, self) or []

        costo = self._pick_cost(rows[0]) if rows else 0.0   # ✅ aquí el cambio real
        self._cost_cache_mat[key] = costo
        return costo



    def _get_costo_cm_laminado(self, lam_tipo_nombre: str | None, lam_marca_nombre: str | None, ancho_cm: float) -> float:
        self._init_cost_maps_if_needed()

        # ✅ si viene ID en vez de nombre
        tid = None
        try:
            if lam_tipo_nombre is not None and str(lam_tipo_nombre).strip().isdigit():
                tid = int(lam_tipo_nombre)
        except Exception:
            tid = None
        if tid is None:
            tid = self._lam_tipo_name_to_id.get(self._norm(lam_tipo_nombre))
        if not tid:
            return 0.0

        # ✅ si viene ID en vez de nombre
        marca_id = None
        try:
            if lam_marca_nombre is not None and str(lam_marca_nombre).strip().isdigit():
                marca_id = int(lam_marca_nombre)
        except Exception:
            marca_id = None
        if marca_id is None and self._norm(lam_marca_nombre):
            marca_id = self._lam_marca_name_to_id.get(self._norm(lam_marca_nombre))

        key = (int(tid), int(marca_id) if marca_id is not None else None, round(float(ancho_cm), 2))
        if key in self._cost_cache_lam:
            return float(self._cost_cache_lam[key] or 0.0)

        params = {"lam_tipo_id": int(tid), "ancho_cm": float(ancho_cm)}
        if marca_id is not None:
            params["lam_marca_id"] = int(marca_id)

        url = "/inventario/laminados/existencias?" + urlencode(params)
        rows = api_get(url, self) or []

        costo = self._pick_cost(rows[0]) if rows else 0.0   # ✅ aquí el cambio real
        self._cost_cache_lam[key] = costo
        return costo



    # -------- UI --------
    def _build(self):
        lay = QVBoxLayout(self)

        # --- Filtros ---
        filter_bar = QWidget()
        filter_bar.setObjectName("FilterBar")
        top = QHBoxLayout(filter_bar)
        top.setContentsMargins(8, 8, 8, 8)
        top.setSpacing(8)

        self.cmb_estado = QComboBox()
        self.cmb_estado.addItems(["Todas", "Finalizadas", "Canceladas"])

        self.dt_desde = QDateEdit(); self.dt_desde.setCalendarPopup(True)
        self.dt_hasta = QDateEdit(); self.dt_hasta.setCalendarPopup(True)
        today = QDate.currentDate()
        self.dt_hasta.setDate(today)
        self.dt_desde.setDate(today.addMonths(-6))
        self.dt_desde.setDisplayFormat("yyyy-MM-dd")
        self.dt_hasta.setDisplayFormat("yyyy-MM-dd")

        self.ed_consec = QLineEdit()
        self.ed_consec.setPlaceholderText("Buscar consecutivo…")

        btn_refrescar = QPushButton("Refrescar");  btn_refrescar.clicked.connect(self._load)
        btn_ajustar   = QPushButton("Ajustar columnas"); btn_ajustar.clicked.connect(lambda: self.tbl.resizeColumnsToContents())
        btn_zoom_mas  = QPushButton("Zoom +");     btn_zoom_mas.clicked.connect(lambda: self._zoom(+1))
        btn_zoom_menos= QPushButton("Zoom −");     btn_zoom_menos.clicked.connect(lambda: self._zoom(-1))
        btn_export    = QPushButton("Exportar a Excel"); btn_export.setProperty("variant", "primary")
        btn_export.clicked.connect(self._export_excel)

        # --- Tabla (CREAR ANTES del menú de columnas) ---
        self.tbl = QTableWidget(0, len(self.HEADERS))
        self.tbl.setHorizontalHeaderLabels(self.HEADERS)
        
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectItems) # Selecciona solo la celda tocada
        self.tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(36)
        self.tbl.setWordWrap(False)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setStretchLastSection(False)

        widths = [
            60,110,140,110,110,120,   # 0..5
            200,95,90,60,             # 6..9
            110,90,90,                # 10..12
            110,110,105,115,          # 13..16
            130,                      # 17
            110,110,110,              # 18..20
            380,70,110                # 21..23
        ]
        for i, w in enumerate(widths):
            if i < self.tbl.columnCount():
                self.tbl.setColumnWidth(i, w)

        # --- Menú de columnas (YA existe self.tbl) ---
        self._build_columns_menu()

        # --- Armar barra superior ---
        top.addWidget(btn_export)
        top.addWidget(QLabel("Estado:")); top.addWidget(self.cmb_estado)
        top.addWidget(QLabel("Desde:"));  top.addWidget(self.dt_desde)
        top.addWidget(QLabel("Hasta:"));  top.addWidget(self.dt_hasta)
        top.addWidget(self.ed_consec, 1)
        top.addStretch(1)
        top.addWidget(btn_zoom_mas); top.addWidget(btn_zoom_menos)
        top.addWidget(btn_ajustar)

        # ✅ Aquí se agrega el botón del menú
        top.addWidget(self.btn_cols)

        top.addWidget(btn_refrescar)

        lay.addWidget(filter_bar)
        lay.addWidget(self.tbl, 1)

        # ✅ caches/mappings para costos
        self._mat_name_to_id = {}
        self._mat_brand_by_matid = {}
        self._lam_tipo_name_to_id = {}
        self._lam_marca_name_to_id = {}
        self._cost_cache_mat = {}
        self._cost_cache_lam = {}


    def _paint_row_bg(self, row: int, bg_hex: str, skip_cols: set[int] | None = None):
        from PySide6.QtGui import QColor
        col = QColor(bg_hex)
        skip = set(skip_cols or [])

        for c in range(self.tbl.columnCount()):
            if c in skip:
                continue

            it = self.tbl.item(row, c)
            if it:
                it.setBackground(col)

            w = self.tbl.cellWidget(row, c)
            if c == 5 and w is not None:
                w.setStyleSheet((w.styleSheet() or "") + f"; background:{bg_hex};")


    def _num(self, v) -> float:
        try:
            return float(v)
        except Exception:
            return 0.0

    def _calc_consumo_esperado(self, r: dict) -> float:
        imp = self._num(r.get("impreso_cm") or r.get("IMPRESO_CM"))
        if imp <= 0:
            imp = self._num(r.get("LARGO") or r.get("largo") or 0)

        rep = int(self._num(r.get("REP") or 0))
        rep = max(rep, 0)

        desp_l = self._num(r.get("desp_largo_cm") or r.get("DESP_LARGO_CM"))
        esp = self._num(r.get("espacio_reps_cm") or 0)

        pasadas = rep + (1 if imp > 0 else 0)
        return max(0.0, (imp * pasadas) + desp_l + (esp * rep))

    # -------- Fetch helpers --------
    def _fetch_finalizadas(self):
        rows = api_get("/ordenes/finalizadas?limite=1000", self)
        if not rows:
            rows = api_get("/ordenes/historial?estado=FINALIZADA&limite=1000", self) or []

        out = []
        for r in rows or []:
            rr = {
                "id_orden": r.get("id_orden"),
                "estado": "FINALIZADA",
                "ruta": r.get("ruta") or r.get("RUTA") or "",
                "FECHA": r.get("FECHA"),
                "fecha_entrega": r.get("fecha_entrega"),
                "CONSECUTIVO": r.get("CONSECUTIVO") or r.get("consecutivo", ""),
                "material": r.get("material"),
                "ancho_rollo_cm": r.get("ANCHO") or r.get("ancho_rollo_cm"),
                "LARGO": r.get("LARGO"),
                "REP": r.get("REP"),
                "impreso_cm": r.get("impreso_cm") or r.get("IMPRESO_CM"),
                "desp_largo_cm": r.get("desp_largo_cm") or r.get("DESP_LARGO_CM"),
                "desp_ancho_cm": r.get("desp_ancho_cm") or r.get("DESP_ANCHO_CM"),
                "lam_tipo": r.get("lam_tipo") or r.get("LAM_TIPO"),
                "lam_marca": r.get("lam_marca") or r.get("LAM_MARCA"),
                "lam_ancho_cm": r.get("lam_ancho_cm") or r.get("LAM_ANCHO_CM"),
                "lam_consumo_cm": (
                    r.get("lam_consumo_cm") or r.get("LAM_CONSUMO_CM") or
                    r.get("lam_consumo") or r.get("LAM_CONSUMO") or
                    r.get("lam_consumo_total") or r.get("LAM_CONSUMO_TOTAL") or
                    r.get("consumo_laminado_cm") or r.get("CONSUMO_LAMINADO_CM") or 0
                ),
                "material_id": r.get("material_id") or r.get("id_material_FINAL"),
                "marca_id": r.get("marca_id") or r.get("fk_id_marca"),
                "marca": r.get("marca") or r.get("MARCA"),

                "merma_cancel_cm": r.get("merma_cancel_cm") or r.get("MERMA_CANCEL_CM"),
                "espacio_reps_cm": r.get("espacio_reps_cm") or 0,
            }
            rr["consumo_esp_cm"] = round(self._calc_consumo_esperado(rr), 2)
            out.append(rr)
        return out

    def _fetch_canceladas(self):
        rows = api_get("/ordenes/canceladas?limite=1000", self) or \
               api_get("/ordenes/historial?estado=CANCELADA&limite=1000", self) or []
        out = []
        for r in rows or []:
            rr = {
                "id_orden": r.get("id_orden"),
                "estado": "CANCELADA",
                "ruta": r.get("ruta") or r.get("RUTA") or "",
                "FECHA": r.get("FECHA"),
                "fecha_entrega": r.get("fecha_entrega"),
                "CONSECUTIVO": r.get("CONSECUTIVO") or r.get("consecutivo", ""),
                "material": r.get("material"),
                "ancho_rollo_cm": r.get("ANCHO") or r.get("ancho_rollo_cm"),
                "LARGO": r.get("LARGO"),
                "REP": r.get("REP"),
                "impreso_cm": r.get("impreso_cm") or r.get("IMPRESO_CM"),
                "desp_largo_cm": r.get("desp_largo_cm") or r.get("DESP_LARGO_CM"),
                "desp_ancho_cm": r.get("desp_ancho_cm") or r.get("DESP_ANCHO_CM"),
                "lam_tipo": r.get("lam_tipo") or r.get("LAM_TIPO"),
                "lam_marca": r.get("lam_marca") or r.get("LAM_MARCA"),
                "lam_ancho_cm": r.get("lam_ancho_cm") or r.get("LAM_ANCHO_CM"),
                "lam_consumo_cm": (
                    r.get("lam_consumo_cm") or r.get("LAM_CONSUMO_CM") or
                    r.get("lam_consumo") or r.get("LAM_CONSUMO") or
                    r.get("lam_consumo_total") or r.get("LAM_CONSUMO_TOTAL") or
                    r.get("consumo_laminado_cm") or r.get("CONSUMO_LAMINADO_CM") or 0
                ),
                "material_id": r.get("material_id") or r.get("id_material_FINAL"),
                "marca_id": r.get("marca_id") or r.get("fk_id_marca"),
                "marca": r.get("marca") or r.get("MARCA"),

                "merma_cancel_cm": r.get("merma_cancel_cm") or r.get("MERMA_CANCEL_CM"),
                "espacio_reps_cm": r.get("espacio_reps_cm") or 0,
            }
            rr["consumo_esp_cm"] = round(self._calc_consumo_esperado(rr), 2)
            out.append(rr)
        return out

    def _load(self):

        self._cost_cache_mat.clear()
        self._cost_cache_lam.clear()
        self.tbl.setRowCount(0)

        estado = self.cmb_estado.currentText()
        if estado == "Finalizadas":
            rows = self._fetch_finalizadas()
        elif estado == "Canceladas":
            rows = self._fetch_canceladas()
        else:
            rows = (self._fetch_finalizadas() or []) + (self._fetch_canceladas() or [])

        rows = rows or []
        rows = self._filtrar_finalizadas_ultimos_6_meses(rows)

        d0 = self.dt_desde.date().toString("yyyy-MM-dd")
        d1 = self.dt_hasta.date().toString("yyyy-MM-dd")
        consec_q = (self.ed_consec.text() or "").strip().lower()

        def fecha_para_filtro(r: dict) -> str:
            return str(r.get("fecha_entrega") or r.get("FECHA") or "").strip()

        def in_range(fecha_str: str) -> bool:
            if not fecha_str:
                return True
            return d0 <= fecha_str <= d1

        filtered = []
        for r in rows:
            if not in_range(fecha_para_filtro(r)):
                continue

            consec_val = str(r.get("CONSECUTIVO", "")).lower()
            if consec_q and consec_q not in consec_val:
                continue

            if "consumo_esp_cm" not in r:
                r["consumo_esp_cm"] = round(self._calc_consumo_esperado(r), 2)

            filtered.append(r)
            self._add_row(r)

        self._rows_cur = filtered

    def _filtrar_finalizadas_ultimos_6_meses(self, rows):
        limite = QDate.currentDate().addMonths(-6)

        out = []
        for r in rows or []:
            estado = (r.get("estado") or "").upper()
            if "FINAL" not in estado:
                out.append(r)
                continue

            ent_str = str(r.get("fecha_entrega") or "").strip()
            d_ent = QDate.fromString(ent_str, "yyyy-MM-dd")
            if d_ent.isValid():
                if d_ent >= limite:
                    out.append(r)
                continue

            f_str = str(r.get("FECHA") or "").strip()
            d_f = QDate.fromString(f_str, "yyyy-MM-dd")
            if (not d_f.isValid()) or (d_f >= limite):
                out.append(r)

        return out

    def _add_row(self, r: dict):
        i = self.tbl.rowCount()
        self.tbl.insertRow(i)

        def set_text(c, val):
            self.tbl.setItem(i, c, QTableWidgetItem("" if val is None else str(val)))

        set_text(0, r.get("id_orden", ""))
        set_text(1, r.get("estado", ""))
        set_text(2, r.get("consumo_esp_cm", ""))
        set_text(3, r.get("FECHA", ""))
        set_text(4, r.get("fecha_entrega", ""))

        ed = QLineEdit(str(r.get("CONSECUTIVO", "")))
        self.tbl.setCellWidget(i, 5, ed)

        set_text(6, r.get("material", ""))
        set_text(7, r.get("ancho_rollo_cm", ""))
        set_text(8, r.get("LARGO", ""))
        set_text(9, r.get("REP", ""))
        set_text(10, r.get("impreso_cm", ""))
        set_text(11, r.get("desp_largo_cm", ""))
        set_text(12, r.get("desp_ancho_cm", ""))

        set_text(13, r.get("lam_tipo", ""))
        set_text(14, r.get("lam_marca", ""))
        set_text(15, r.get("lam_ancho_cm", ""))
        set_text(16, r.get("lam_consumo_cm", ""))
        set_text(17, r.get("merma_cancel_cm", ""))

        # ===== ✅ CÁLCULO DE COSTOS =====
        consumo_esp = self._safe_float(r.get("consumo_esp_cm"))
        ancho_roll = self._safe_float(r.get("ancho_rollo_cm"))

        mat_id = r.get("material_id")
        marca_id = r.get("marca_id")

        costo_cm_mat = self._get_costo_cm_material_ids(mat_id, marca_id, ancho_roll)
        costo_mat = consumo_esp * ancho_roll * costo_cm_mat

        costo_mat = consumo_esp * ancho_roll * costo_cm_mat

        lam_consumo = self._safe_float(r.get("lam_consumo_cm"))
        lam_ancho = self._safe_float(r.get("lam_ancho_cm")) or ancho_roll
        costo_cm_lam = self._get_costo_cm_laminado(r.get("lam_tipo"), r.get("lam_marca"), lam_ancho)
        costo_lam = lam_consumo * lam_ancho * costo_cm_lam

        costo_total = costo_mat + costo_lam

        self._set_cost_cell(i, self.COL_COSTO_MAT, costo_mat)
        self._set_cost_cell(i, self.COL_COSTO_LAM, costo_lam)
        self._set_cost_cell(i, self.COL_COSTO_TOTAL, costo_total)


        set_text(21, r.get("ruta", ""))

        btn_open = QPushButton("Abrir")
        ruta_txt = r.get("ruta", "")
        btn_open.clicked.connect(lambda _=False, p=ruta_txt: self._abrir_ruta(p))
        self.tbl.setCellWidget(i, 22, btn_open)

        btn = QPushButton("Guardar")
        btn.clicked.connect(partial(self._save_consec, i, int(r.get("id_orden", 0) or 0)))
        self.tbl.setCellWidget(i, 23, btn)

        st = (r.get("estado", "") or "").upper()
        skip = set(self.COST_COLS)

        if "PENDIENTE" in st:
            self._paint_row_bg(i, PENDIENTE_BG, skip_cols=skip)
        elif "EN PROCESO" in st:
            self._paint_row_bg(i, EN_PROCESO_BG, skip_cols=skip)
        elif "FINAL" in st:
            self._paint_row_bg(i, FINALIZADA_BG, skip_cols=skip)
        elif "CANCEL" in st:
            self._paint_row_bg(i, CANCELADA_BG, skip_cols=skip)


    def _save_consec(self, row: int, oid: int):
        w = self.tbl.cellWidget(row, 5)
        consec = w.text().strip() if isinstance(w, QLineEdit) else ""
        if not oid or not consec:
            QMessageBox.warning(self, "Error", "ID y Consecutivo son obligatorios.")
            return

        payload = {"id_orden": oid, "consecutivo": consec}
        r = api_post("/ordenes/editar-consecutivo", payload, self)
        if r:
            QMessageBox.information(self, "OK", "Consecutivo actualizado.")

    # -------- Exportación a Excel (solo columnas visibles) --------
    def _export_excel(self):
        if self.tbl.rowCount() == 0:
            QMessageBox.information(self, "Exportar", "No hay filas para exportar (aplica filtros y vuelve a intentar).")
            return

        visible_cols = [c for c in range(self.tbl.columnCount()) if not self.tbl.isColumnHidden(c)]
        if not visible_cols:
            QMessageBox.information(self, "Exportar", "No hay columnas visibles para exportar.")
            return

        default_name = f"reporte_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Exportar a Excel", default_name, "Excel (*.xlsx)")
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            QMessageBox.warning(self, "Falta dependencia", "Instala openpyxl:\n\npip install openpyxl")
            return

        def header_text(col: int) -> str:
            it = self.tbl.horizontalHeaderItem(col)
            return it.text() if it else (self.HEADERS[col] if col < len(self.HEADERS) else f"Col{col}")

        def cell_text(row: int, col: int) -> str:
            w = self.tbl.cellWidget(row, col)
            if w is not None:
                if hasattr(w, "text"):
                    try:
                        return str(w.text() or "")
                    except Exception:
                        return ""
                return ""

            it = self.tbl.item(row, col)
            if not it:
                return ""

            # ✅ Si hay valor numérico crudo (como en costos), úsalo
            try:
                from PySide6.QtCore import Qt
                raw = it.data(Qt.UserRole)
                if isinstance(raw, (int, float)):
                    return str(raw)
            except Exception:
                pass

            return it.text() if it else ""


        def to_number(s: str):
            s = (s or "").strip()
            if not s:
                return None
            t = s.replace(".", "").replace(",", ".") if (s.count(",") == 1 and s.count(".") >= 1) else s.replace(",", ".")
            try:
                return float(t)
            except Exception:
                return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        head_fill = PatternFill("solid", fgColor="00897B")
        head_font = Font(bold=True, color="FFFFFF")
        center = Alignment(vertical="center")
        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for out_c, col in enumerate(visible_cols, start=1):
            title = header_text(col)
            cell = ws.cell(row=1, column=out_c, value=title)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = center
            cell.border = border

        nrows = self.tbl.rowCount()
        for r in range(nrows):
            for out_c, col in enumerate(visible_cols, start=1):
                txt = cell_text(r, col)
                num = to_number(txt)
                if num is not None and txt.strip() not in {"-", ""}:
                    ws.cell(row=r + 2, column=out_c, value=num).border = border
                else:
                    ws.cell(row=r + 2, column=out_c, value=txt).border = border

        def is_date_col(title: str) -> bool:
            t = (title or "").strip().lower()
            return t in {"fecha", "entrega"}

        def is_money_col(title: str) -> bool:
            t = (title or "").strip().lower()
            return "costo" in t

        def is_numeric_cm_col(title: str) -> bool:
            t = (title or "").strip().lower()
            if "(cm)" in t:
                return True
            if t in {"ancho roll", "largo", "rep"}:
                return True
            if t.startswith("lam.") and ("ancho" in t or "consumo" in t):
                return True
            if "consumo esp" in t:
                return True
            return False

        for out_c, col in enumerate(visible_cols, start=1):
            title = header_text(col)
            if is_date_col(title):
                for rr in range(2, nrows + 2):
                    ws.cell(row=rr, column=out_c).number_format = "yyyy-mm-dd"
            elif is_money_col(title):
                for rr in range(2, nrows + 2):
                    ws.cell(row=rr, column=out_c).number_format = '$#,##0.##'
            elif is_numeric_cm_col(title):
                for rr in range(2, nrows + 2):
                    ws.cell(row=rr, column=out_c).number_format = "0.00"

        for out_c, col in enumerate(visible_cols, start=1):
            px = self.tbl.columnWidth(col)
            ws.column_dimensions[get_column_letter(out_c)].width = max(8, min(60, int(px / 7)))

        ws.freeze_panes = "A2"
        last_cell = f"{get_column_letter(len(visible_cols))}{nrows + 1}"
        try:
            t = Table(displayName="ReporteGF", ref=f"A1:{last_cell}")
            t.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(t)
        except Exception:
            pass

        try:
            wb.save(path)
            QMessageBox.information(self, "Exportar", f"Archivo guardado:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error al guardar", str(e))


# =========================
#     IMPRESIÓN (tabla)
# =========================

# =========================
#     IMPRESIÓN (tabla)
# =========================
class HScrollTable(QTableWidget):
    """Tabla que permite scroll horizontal con Shift + rueda."""
    def wheelEvent(self, e):
        if e.modifiers() & Qt.ShiftModifier:
            sb = self.horizontalScrollBar()
            sb.setValue(sb.value() - e.angleDelta().y())
            e.accept()
        else:
            super().wheelEvent(e)


class PagImpresion(QWidget):
    # Nota: quitamos la columna "Reps" editable y añadimos "Ruta"
    HEADERS = [
        "ID","Estado","Fecha","Entrega","Material","ancho roll","Largo","Rep",
        "Desp.L(cm)","Desp.A(cm)",
        "Lam.tipo","Lam.marca","Lam.ancho(cm)","Lam.consumo(cm)",
        "Merma cancel(cm)","Esp.reps(cm)","Observaciones","Ruta",
        "Consumo esp.(cm)",   # ✅ NUEVA
        "Acciones"
    ]

    def __init__(self):
        super().__init__()
        
        # 1. Inicializar variables de estado
        self._lam_tipos   = []
        self._lam_marcas  = []
        self._lam_medidas = []
        self._all_materials_list = [] # Guardará la lista completa de materiales del backend
        self._all_widths_list = []

        # 2. Configuración de persistencia (QSettings) ANTES de construir la UI
        self.settings = QSettings("AOP", "ControlProduccion")
        self._colores_recientes = self.settings.value("impresion/colores_recientes", [])
        
        # Asegurarnos de limitar la lista (ej. últimos 10 colores)
        if len(self._colores_recientes) > 10:
            self._colores_recientes = self._colores_recientes[:10]

        # 3. Construir la interfaz gráfica UNA SOLA VEZ
        self._build()

        # 4. Definir los índices de columnas (según HEADERS)
        self.COL_LAM_CONS    = self.HEADERS.index("Lam.consumo(cm)")
        self.COL_LAM_ANCHO   = self.HEADERS.index("Lam.ancho(cm)")
        self.COL_LAM_MARCA   = self.HEADERS.index("Lam.marca")
        self.COL_LAM_TIPO    = self.HEADERS.index("Lam.tipo")
        self.COL_CONSUMO_ESP = self.HEADERS.index("Consumo esp.(cm)")
        self.COL_LARGO       = self.HEADERS.index("Largo")
        self.COL_REP         = self.HEADERS.index("Rep")
        self.COL_DESP_L      = self.HEADERS.index("Desp.L(cm)")
        self.COL_DESP_A      = self.HEADERS.index("Desp.A(cm)")
        self.COL_ESP_REPS    = self.HEADERS.index("Esp.reps(cm)")
        self.COL_ANCHO_ROLL  = self.HEADERS.index("ancho roll")

        # 5. Cargas iniciales desde el backend UNA SOLA VEZ
        self._load_lam_catalogs()
        self._load_filter_catalogs() 
        self._load_panel()



    # ---------- UI ----------
    def _build(self):
        lay = QVBoxLayout(self)

        # --- Barra superior ---
        top = QHBoxLayout()
        self.chk_incluir_fin = QCheckBox("Incluir FINALIZADAS en la tabla")

        self.chk_incluir_fin.stateChanged.connect(self._load_panel)
        self.chk_incluir_fin.setToolTip("Si está activo, muestra FINALIZADAS sólo de los últimos 14 días.")
        self.chk_incluir_fin.setStyleSheet("""
        QCheckBox::indicator { width: 18px; height: 18px; }
        QCheckBox::indicator:checked {
            background: #4CAF50; border: 1px solid #4CAF50;
        }
        QCheckBox::indicator:unchecked {
            background: white; border: 1px solid #BDBDBD;
        }
        """)

        btn_refresh = QPushButton("Refrescar");         btn_refresh.clicked.connect(self._load_panel)
        btn_clean   = QPushButton("Limpiar finalizadas");btn_clean.clicked.connect(self._limpiar_finalizadas)
        self.btn_zoom_in  = QPushButton("Zoom +");       self.btn_zoom_in.clicked.connect(lambda: self._zoom(+1))
        self.btn_zoom_out = QPushButton("Zoom −");       self.btn_zoom_out.clicked.connect(lambda: self._zoom(-1))
        self.btn_fit      = QPushButton("Ajustar columnas"); self.btn_fit.clicked.connect(self._fit_columns)
        self.btn_zoom_in.setProperty("variant", "secondary")
        self.btn_zoom_out.setProperty("variant", "secondary")
        self.btn_fit.setProperty("variant", "secondary")
        top.addWidget(self.chk_incluir_fin); top.addStretch(1)
        top.addWidget(self.btn_zoom_in); top.addWidget(self.btn_zoom_out)
        top.addWidget(self.btn_fit); top.addWidget(btn_refresh); top.addWidget(btn_clean)

        # --- Tabla (crear primero, luego configurar) ---
        self.tbl = HScrollTable(0, len(self.HEADERS))
        f = self.tbl.font()
        f.setPointSize(max(9, f.pointSize() - 2))  # baja ~2pt
        self.tbl.setFont(f)
        self.tbl.verticalHeader().setDefaultSectionSize(max(22, f.pointSize() + 12))
        self.tbl.setStyleSheet("""
        QHeaderView::section { padding: 6px 8px; }
        QTableWidget { 
            gridline-color: #E0E0E0; 
            outline: none; /* Quita el recuadro punteado feo de Windows */
            selection-background-color: rgba(0, 0, 0, 20); /* Tinte oscuro al 8% de opacidad para NO borrar el color base */
            selection-color: black;
        }
        QTableWidget::item:selected {
            border: 2px solid #2196F3; /* Borde azul moderno solo para la celda seleccionada */
        }
        """)
        
        self.tbl.setHorizontalHeaderLabels(self.HEADERS)
        self.tbl.setAlternatingRowColors(False)
        
        # 👇 CORRECCIÓN CLAVE: Usar SelectItems en lugar de SelectRows 👇
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tbl.setWordWrap(False)
        self.tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tbl.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.tbl.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.tbl.verticalHeader().setDefaultSectionSize(36)  # altura consistente
        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._show_context_menu)
        hdr = self.tbl.horizontalHeader()
        hdr.setStretchLastSection(False)
        hdr.setSectionResizeMode(QHeaderView.Interactive)

        # --- Filtros (solo una vez; quitar el bloque duplicado) ---
        # --- NUEVOS FILTROS DESDE 0 (Checklists) ---
        filter_bar = QWidget()
        filter_bar.setObjectName("FilterBar")
        flt = QHBoxLayout(filter_bar)
        flt.setContentsMargins(8, 8, 8, 8)
        flt.setSpacing(8)

        # 1. Rango de Fechas (se mantiene igual)
        self.fil_ent_desde = QDateEdit()
        self.fil_ent_desde.setCalendarPopup(True)
        self.fil_ent_desde.setDisplayFormat("yyyy-MM-dd")
        self.fil_ent_hasta = QDateEdit()
        self.fil_ent_hasta.setCalendarPopup(True)
        self.fil_ent_hasta.setDisplayFormat("yyyy-MM-dd")

        # 2. TIPO MATERIAL (Categoría) - Checklist fijo
        self.fil_cat = CheckableComboBox("Categoría...")
        self.fil_cat.addItem("ADHESIVO", 1) # ID 1 para Adhesivo
        self.fil_cat.addItem("LONA", 0)     # ID 0 para Lona

        # 3. MATERIAL - Checklist dinámico
        self.fil_mat = CheckableComboBox("Material...")

        # 4. Medida rollo (Ancho) - Checklist dinámico
        self.fil_medida = CheckableComboBox("Medida...")

        btn_reset_filtros = QPushButton("Borrar filtros")
        btn_reset_filtros.setProperty("variant", "secondary")
        btn_reset_filtros.clicked.connect(self._reset_filters)

        # Conexiones de Señales 👇
        
        # Fechas (al vuelo)
        self.fil_ent_desde.dateChanged.connect(self._apply_filters)
        self.fil_ent_hasta.dateChanged.connect(self._apply_filters)
        
        # Filtros dependientes: Al cambiar Categoría, actualiza Material y Medida
        self.fil_cat.selectionChanged.connect(self._on_fil_category_changed)
        self.fil_cat.selectionChanged.connect(self._apply_filters) # También filtra
        
        # Material y Medida (al vuelo cuando se cierra el popup)
        self.fil_mat.selectionChanged.connect(self._apply_filters)
        self.fil_medida.selectionChanged.connect(self._apply_filters)

        # Organización en el Layout 👇
        flt.addWidget(QLabel("Entrega desde:"))
        flt.addWidget(self.fil_ent_desde)
        flt.addWidget(QLabel("hasta:"))
        flt.addWidget(self.fil_ent_hasta)
        
        flt.addSpacing(12)
        flt.addWidget(QLabel("Tipo Material:"))
        flt.addWidget(self.fil_cat, 1) # Estira un poco
        flt.addSpacing(6)
        flt.addWidget(QLabel("Material:"))
        flt.addWidget(self.fil_mat, 2) # Estira más
        flt.addSpacing(6)
        flt.addWidget(QLabel("Medida rollo (cm):"))
        flt.addWidget(self.fil_medida, 1) # Estira un poco
        
        flt.addStretch(1)
        flt.addWidget(btn_reset_filtros)

        lay.addLayout(top)
        lay.addWidget(filter_bar)   # ⬅️ suficiente; no dupliques el bloque de filtros
        # --- Anchos (19 columnas; coincide con HEADERS) ---
        widths = [60,110,110,110,200,90,90,60,110,110,210,110,110,115,125,110,220,320,140,220]
        for i, w in enumerate(widths):
            self.tbl.setColumnWidth(i, w)

        lay.addWidget(self.tbl, 1)
        self._reset_filters()


    def _paint_row_bg(self, row: int, bg_hex: str):
        from PySide6.QtGui import QColor
        col = QColor(bg_hex)
        # Celdas con texto
        for c in range(self.tbl.columnCount()):
            it = self.tbl.item(row, c)
            if it:
                it.setBackground(col)
            # Celdas con widgets (spin/combos/botones)
            w = self.tbl.cellWidget(row, c)
            if w:
                w.setStyleSheet(w.styleSheet() + f" ;background:{bg_hex};")

    

        # --- helpers de colores/estilos por estado ---
    def _estado_colors(self, estado_txt: str) -> tuple[str, str]:
        """
        Devuelve (bg_hex, fg_hex_sugerido) para el estado.
        """
        e = (estado_txt or "").upper()
        if "FINAL" in e:
            return FINALIZADA_BG, "#1B5E20"     # verde/fuente verde oscuro
        if "CANCEL" in e:
            return CANCELADA_BG, "#7F0000"
        if "PROCESO" in e:
            return EN_PROCESO_BG, "#4E3B00"     # amarillo/ámbar
        # default: PENDIENTE
        return PENDIENTE_BG, "#5D4037"

    def _tintar_columna_estado_y_acciones(self, row: int, estado_txt: str):
        """
        Aplica el color de estado en:
          - Columna 'Estado' (col 1) -> background del QTableWidgetItem
          - Columna 'Acciones' (col 18) -> background del QWidget contenedor
        """
        from PySide6.QtGui import QColor, QBrush
        bg, _fg = self._estado_colors(estado_txt)

        # Columna "Estado"
        it = self.tbl.item(row, 1)
        if it:
            it.setBackground(QColor(bg))

        # Columna "Acciones": es un QWidget, lo estilamos vía stylesheet
        w = self.tbl.cellWidget(row, 18)
        if w:
            # fondo suave + borde redondeado para que se vea como "pastilla"
            w.setStyleSheet(
                (w.styleSheet() or "")
                + f" QWidget{{background:{bg}; border-radius:10px;}} "
            )





    def _fit_columns(self):
        self.tbl.resizeColumnsToContents()

    def _zoom(self, step):
        f = self.tbl.font()
        f.setPointSize(max(8, f.pointSize() + step))
        self.tbl.setFont(f)
        self.tbl.verticalHeader().setDefaultSectionSize(f.pointSize() + 14)


    def _safe_float_txt(self, v) -> float:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return 0.0

    def _calc_consumo_esperado_live(self, row: int) -> float:
        # Largo (col 6) y Rep (col 7) son items de texto
        largo = self._safe_float_txt(self.tbl.item(row, 6).text() if self.tbl.item(row, 6) else 0)
        rep   = int(self._safe_float_txt(self.tbl.item(row, 7).text() if self.tbl.item(row, 7) else 0))
        rep = max(rep, 0)

        # Desp.L (col 8) y Esp.reps (col 15) son widgets
        w_desp_l = self.tbl.cellWidget(row, 8)
        desp_l = float(w_desp_l.value()) if isinstance(w_desp_l, QDoubleSpinBox) else 0.0

        w_esp = self.tbl.cellWidget(row, 15)
        esp = float(w_esp.value()) if isinstance(w_esp, QDoubleSpinBox) else 0.0

        # ✅ política correcta: REP = repeticiones adicionales
        return max(0.0, (largo * (rep + 1)) + desp_l + (esp * rep))

    def _update_consumo_cell(self, row: int):
        col_consumo = self.COL_CONSUMO_ESP
        val = self._calc_consumo_esperado_live(row)

        it = self.tbl.item(row, col_consumo)
        if it is None:
            it = QTableWidgetItem("")
            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
            self.tbl.setItem(row, col_consumo, it)

        it.setText(f"{val:.2f}")

        # --- NUEVA AUTOMATIZACIÓN DE LAMINADO ---
        # 1. Verificamos el estado para no alterar el historial de órdenes ya cerradas
        it_estado = self.tbl.item(row, 1) # Columna de Estado
        estado_txt = (it_estado.text() if it_estado else "").upper()
        
        if "FINAL" not in estado_txt and "CANCEL" not in estado_txt:
            
            # 2. Leemos si es SIN LAMINAR
            w_tipo = self.tbl.cellWidget(row, self.COL_LAM_TIPO)
            lam_tipo_txt = (w_tipo.currentText() if isinstance(w_tipo, QComboBox) else "").strip().upper()
            
            # 3. Obtenemos la casilla de consumo de laminado
            sb_lam_cons = self.tbl.cellWidget(row, self.COL_LAM_CONS)
            
            if isinstance(sb_lam_cons, QDoubleSpinBox):
                # Aplicamos la regla lógica:
                if lam_tipo_txt == "SIN LAMINAR" or not lam_tipo_txt:
                    sb_lam_cons.setValue(0.0)
                else:
                    sb_lam_cons.setValue(val + 15.0)




    def _guardar_esp_reps(self, id_orden: int, valor: float):
        try:
            r = requests.post(
                f"{API}/ordenes/set-espacio-reps",
                json={"id_orden": int(id_orden), "espacio_reps_cm": float(valor)},
                timeout=TIMEOUT
            )
            r.raise_for_status()
        except Exception as e:
            _show_err(self, e)


    # ---------- catálogos lam ----------
    def _load_lam_catalogs(self):
        # Tipos
        raw_tipos = api_get("/catalogos/laminados/tipos", self) or []
        self._lam_tipos = [{
            "id":     t.get("id") or t.get("id_laminado") or t.get("id_tipo") or t.get("laminado_id"),
            "nombre": t.get("nombre") or t.get("nombre_laminado") or t.get("laminado") or t.get("tipo"),
        } for t in raw_tipos]

        # Marcas (sin cambios)
        self._lam_marcas  = api_get("/catalogos/laminados/marcas", self)  or []
        # Medidas (sin cambios)
        self._lam_medidas = api_get("/catalogos/laminados/medidas", self) or []


    def _id_sin_laminado(self):
        """Devuelve el id del catálogo cuyo nombre es 'SIN LAMINAR' (o None si no existe)."""
        for t in self._lam_tipos:
            if (t.get("nombre") or "").strip().upper() == "SIN LAMINAR":
                return t.get("id")
        return None


    # ---------- data ----------
    def _fetch_panel(self):
        include_fin = 1 if self.chk_incluir_fin.isChecked() else 0
        return api_get(f"/ordenes/panel_impresion?incluir_finalizadas={include_fin}", self) or []


    def _filtrar_finalizadas_ultimas_2_semanas(self, rows):
        """
        Si la casilla está activa, deja FINALIZADAS con FINALIZADO_AT >= hoy-14d.
        Fallback: fecha_entrega, y si no, FECHA.
        """
        if not self.chk_incluir_fin.isChecked():
            return rows

        limite = QDate.currentDate().addDays(-14)

        def _to_qdate(v) -> QDate | None:
            s = (str(v or "").strip())
            if not s:
                return None
            # Soporta "YYYY-MM-DD" y también "YYYY-MM-DD HH:MM:SS" / ISO
            s = s[:10]
            d = QDate.fromString(s, "yyyy-MM-dd")
            return d if d.isValid() else None

        out = []
        for r in rows or []:
            estado = (r.get("estado") or "").upper()
            if "FINAL" not in estado:
                out.append(r)
                continue

            d = _to_qdate(r.get("finalizado_at")) or _to_qdate(r.get("fecha_entrega")) or _to_qdate(r.get("FECHA"))
            if (d is None) or (d >= limite):
                out.append(r)

        return out



    # ---------- abrir ruta ----------
    def _abrir_ruta(self, ruta_local: str | None, ruta_http: str | None = None):
        """
        Abre el arte asociado a la orden.
        Prioridad:
        1) Si 'ruta_local' ya es http(s), abrirla como URL;
        2) Intentar abrir como archivo local/UNC;
        3) Fallback: abrir 'ruta_http' si llega desde backend.
        """
        if not ruta_local and not ruta_http:
            QMessageBox.warning(self, "Archivo", "No hay RUTA asociada a la orden.")
            return
        if ruta_local and str(ruta_local).lower().startswith(("http://","https://")):
            QDesktopServices.openUrl(QUrl(ruta_local)); return
        if ruta_local and QDesktopServices.openUrl(QUrl.fromLocalFile(ruta_local)):
            return
        if ruta_http:
            QDesktopServices.openUrl(QUrl(ruta_http)); return
        QMessageBox.warning(self, "Archivo", "No se pudo abrir la ruta indicada.")

    # ---------- cargar tabla ----------
    def _load_panel(self):
        # --- 1. CAPTURAR EL ESTADO (MEMORIA UI) ANTES DE RECARGAR ---
        estado_previo = {}
        for r in range(self.tbl.rowCount()):
            it_id = self.tbl.item(r, 0)
            if not it_id: continue
            try:
                oid = int(it_id.text())
                
                # Guardamos índices de combos y valores de los spinboxes
                cb_marca = self.tbl.cellWidget(r, self.COL_LAM_MARCA)
                cb_ancho = self.tbl.cellWidget(r, self.COL_LAM_ANCHO)
                
                estado_previo[oid] = {
                    "desp_l": self._g(r, self.COL_DESP_L, float),
                    "desp_a": self._g(r, self.COL_DESP_A, float),
                    "lam_cons": self._g(r, self.COL_LAM_CONS, float),
                    "merma": self._g(r, 14, float), # 14 = Merma cancel
                    "marca_idx": cb_marca.currentIndex() if isinstance(cb_marca, QComboBox) else -1,
                    "ancho_idx": cb_ancho.currentIndex() if isinstance(cb_ancho, QComboBox) else -1,
                }
            except Exception:
                pass

        # --- 2. LIMPIAR Y RECARGAR DESDE BASE DE DATOS ---
        self.tbl.setRowCount(0)

        rows = self._fetch_panel()
        rows = self._filtrar_finalizadas_ultimas_2_semanas(rows)

        for row in (rows or []):
            self._add_row(row)

        # --- 3. RESTAURAR LOS VALORES TECLEADOS POR EL USUARIO ---
        for r in range(self.tbl.rowCount()):
            it_id = self.tbl.item(r, 0)
            if not it_id: continue
            try:
                oid = int(it_id.text())
                prev = estado_previo.get(oid)
                
                if prev:
                    # Restaurar Desp.L
                    if prev["desp_l"] > 0:
                        w = self.tbl.cellWidget(r, self.COL_DESP_L)
                        if isinstance(w, QDoubleSpinBox): w.setValue(prev["desp_l"])
                    
                    # Restaurar Desp.A (solo si es editable manualmente)
                    if prev["desp_a"] > 0:
                        w = self.tbl.cellWidget(r, self.COL_DESP_A)
                        if isinstance(w, QDoubleSpinBox) and w.isEnabled(): 
                            w.setValue(prev["desp_a"])
                            
                    # Restaurar Lam.consumo (si es editable)
                    if prev["lam_cons"] > 0:
                        w = self.tbl.cellWidget(r, self.COL_LAM_CONS)
                        if isinstance(w, QDoubleSpinBox) and w.isEnabled(): 
                            w.setValue(prev["lam_cons"])
                            
                    # Restaurar Merma cancel
                    if prev["merma"] > 0:
                        w = self.tbl.cellWidget(r, 14)
                        if isinstance(w, QDoubleSpinBox): w.setValue(prev["merma"])
                        
                    # Restaurar combos de laminado (Marca y Ancho)
                    if prev["marca_idx"] >= 0:
                        cb = self.tbl.cellWidget(r, self.COL_LAM_MARCA)
                        if isinstance(cb, QComboBox) and cb.isEnabled(): 
                            cb.setCurrentIndex(prev["marca_idx"])
                        
                    if prev["ancho_idx"] >= 0:
                        cb = self.tbl.cellWidget(r, self.COL_LAM_ANCHO)
                        if isinstance(cb, QComboBox) and cb.isEnabled(): 
                            cb.setCurrentIndex(prev["ancho_idx"])
                            
            except Exception:
                pass

        # Re-aplicar filtros tras recargar la tabla
        self._apply_filters()


    


    #logica de filtrado:
        # ---------- Filtros ----------
    def _reset_filters(self):
        # Rango de fechas “abierto” (se mantiene lógica de fechas que tenías o la de 15 días)
        hoy = QDate.currentDate()
        self.fil_ent_desde.setDate(hoy.addDays(-15))
        self.fil_ent_hasta.setDate(hoy.addDays(30)) # Warning de UX, mejor mirar futuro
        
        # ✅ NUEVA LÓGICA DE RESET PARA CHECKLISTS 👇
        self.fil_cat.checkAll(False) # Deschequear todo en Categoría
        
        # Forzar repoblado de dependientes (carga todos los materiales)
        self._poblar_filtros_dependientes() 
        
        self.fil_mat.checkAll(False) # Deschequear todo en Material
        self.fil_medida.checkAll(False) # Deschequear todo en Medida
        
        self._apply_filters()

    def _cell_date(self, row: int, col: int) -> QDate | None:
        it = self.tbl.item(row, col)
        s = (it.text().strip() if it else "")
        d = QDate.fromString(s, "yyyy-MM-dd")
        return d if d.isValid() else None

    def _cell_str(self, row: int, col: int) -> str:
        it = self.tbl.item(row, col)
        return (it.text() if it else "").strip()

    def _cell_float(self, row: int, col: int) -> float | None:
        it = self.tbl.item(row, col)
        if not it:
            return None
        try:
            return float(it.text())
        except Exception:
            return None

    # ---------- helpers de fecha/alertas ----------
    def _parse_entrega_qdate(self, entrega_str: str) -> QDate | None:
        s = (entrega_str or "").strip()
        d = QDate.fromString(s, "yyyy-MM-dd")
        return d if d.isValid() else None

    def _is_due_soon(self, entrega: QDate) -> bool:
        """Entrega HOY o MAÑANA."""
        hoy = QDate.currentDate()
        return entrega == hoy or entrega == hoy.addDays(1)

    def _is_overdue(self, entrega: QDate) -> bool:
        """Entrega vencida (cualquier día antes de HOY)."""
        hoy = QDate.currentDate()
        return entrega < hoy

    def _style_entrega_cell(self, row: int, entrega_str: str):
        """
        Aplica:
        - ⚠️ si entrega es HOY o MAÑANA
        - Fondo naranja si entrega está vencida (entrega < hoy)
        SOLO afecta la celda 'Entrega' (col 3).
        """
        from PySide6.QtGui import QColor

        col_entrega = 3  # "Entrega"
        d = self._parse_entrega_qdate(entrega_str)

        it = self.tbl.item(row, col_entrega)
        if not it:
            return

        # Reset básico (por si se recarga)
        it.setBackground(QColor("white"))

        if not d:
            # si no se parsea, dejamos tal cual
            it.setText(str(entrega_str or ""))
            return

        if self._is_overdue(d):
            # vencida => celda naranja
            it.setText(d.toString("yyyy-MM-dd"))
            it.setBackground(QColor("#FFA500"))  # naranja
            it.setToolTip("ENTREGA VENCIDA")
            return

        if self._is_due_soon(d):
            # hoy/mañana => emoji ⚠️
            it.setText(f"{d.toString('yyyy-MM-dd')} ⚠️")
            it.setToolTip("ENTREGA HOY o MAÑANA")
            return

        # normal
        it.setText(d.toString("yyyy-MM-dd"))
        it.setToolTip("")




    def _apply_filters(self):
        """Oculta/muestra filas combinando Fechas (rango) y Checklists (coincidencia en lista)."""
        # Obtenemos rango de fechas
        d0 = self.fil_ent_desde.date()
        d1 = self.fil_ent_hasta.date()
        
        # ✅ NUEVOS GETTERS DE DATOS CHEQUEADOS 👇
        cats_checked = self.fil_cat.checkedData() # Lista de IDs [0, 1]
        mats_checked = self.fil_mat.checkedData() # Lista de Material IDs
        medidas_checked = self.fil_medida.checkedData() # Lista de anchos (floats)

        # Banderas para optimizar (si no hay nada chequeado, no filtramos por ese campo)
        filter_cat = len(cats_checked) > 0
        filter_mat = len(mats_checked) > 0
        filter_medida = len(medidas_checked) > 0

        # Bloqueamos señales de la tabla para velocidad
        self.tbl.setUpdatesEnabled(False)
        try:
            for r in range(self.tbl.rowCount()):
                ok = True

                # 1. Filtro Fecha Entrega -> columna 3
                dent = self._cell_date(r, 3)
                if dent is not None:
                    if dent < d0 or dent > d1:
                        ok = False

                # 2. Filtro Categoría (Dato Oculto en Columna 0 ID) 👇
                if ok and filter_cat:
                    it_id = self.tbl.item(r, 0)
                    row_cat_id = it_id.data(Qt.UserRole) if it_id else None
                    if row_cat_id not in cats_checked:
                        ok = False

                # 3. Filtro Material ID (Dato Oculto en Columna 4 Material) 👇
                if ok and filter_mat:
                    it_mat = self.tbl.item(r, 4)
                    row_mat_id = it_mat.data(Qt.UserRole) if it_mat else None
                    if row_mat_id not in mats_checked:
                        ok = False

                # 4. Filtro Medida Rollo -> columna 5 👇
                if ok and filter_medida:
                    try:
                        it_ancho = self.tbl.item(r, 5)
                        row_width = float(it_ancho.text()) if it_ancho else None
                        # Comparamos floats con tolerancia sutil
                        match = False
                        for w in medidas_checked:
                            if abs(w - row_width) < 0.01:
                                match = True; break
                        if not match:
                            ok = False
                    except Exception:
                        ok = False # Si no es numérico, lo ocultamos si hay filtro activo

                self.tbl.setRowHidden(r, not ok)
                
        finally:
            # Desbloqueamos señales
            self.tbl.setUpdatesEnabled(True)


    # ---------- catálogos para filtros ----------
    # ---------- catálogos para filtros ----------
    def _load_filter_catalogs(self):
        """Carga inicial de materiales y medidas desde el backend para los filtros."""
        # 1. Obtenemos TODOS los materiales
        self._all_materials_list = api_get("/catalogos/materiales", self) or []
        
        # 2. Obtenemos las medidas iterando sobre cada material 
        # (El backend exige saber de qué material queremos las medidas para no dar error 404)
        seen = set()
        self._all_widths_list = []
        
        for mat in self._all_materials_list:
            mat_id = mat.get("id")
            if mat_id:
                # Pedimos las medidas específicas de este material al servidor
                raw_medidas = api_get(f"/catalogos/medidas?material_id={mat_id}", self) or []
                
                # Las extraemos y filtramos para no tener medidas repetidas en el checklist
                # (Ej. si tanto el Adhesivo como la Lona tienen rollo de 152.0, solo aparecerá una vez)
                for m in raw_medidas:
                    try:
                        w = float(m["ancho"])
                        if w not in seen:
                            seen.add(w)
                            self._all_widths_list.append(w)
                    except Exception:
                        pass
                        
        # 3. Ordenamos las medidas de menor a mayor (ej. 105.0, 122.0, 152.0, 320.0)
        self._all_widths_list.sort() 

        # 4. Poblado inicial de los filtros desplegables
        self._poblar_filtros_dependientes()

    def _on_fil_category_changed(self):
        """Se dispara cuando cambia la selección de Categoría (Lona/Adhesivo)."""
        # Repoblamos Material y Medida basándonos en los checks actuales de Categoría
        self._poblar_filtros_dependientes()

    def _poblar_filtros_dependientes(self):
        """Filtra y puebla los combos de Material y Medida según la Categoría seleccionada."""
        # 1. Obtener qué categorías están chequeadas
        checked_cats = self.fil_cat.checkedData() # Lista de IDs [0, 1]

        # Si no hay nada chequeado, asumimos "TODAS"
        show_all = len(checked_cats) == 0

        # Memorizar selecciones actuales para intentar restaurarlas tras el repoblado
        prev_mats = self.fil_mat.checkedData()
        prev_medidas = self.fil_medida.checkedData()

        # --- Repoblar MATERIAL ---
        self.fil_mat.clear()
        
        # Heurística para vincular material a categoría si el backend no envía cat_id
        def es_adhesivo_nombre(nombre):
            n = (nombre or "").upper()
            return any(tok in n for tok in ("ADH", "ADHESIVO", "VINIL", "VINILO", "STICKER"))

        for m in self._all_materials_list:
            mat_is_adh = es_adhesivo_nombre(m["nombre"])
            # Mat es adhesivo (1) o lona (0)
            mat_cat = 1 if mat_is_adh else 0
            
            # Si show_all o la categoría del material está chequeada, lo añadimos
            if show_all or mat_cat in checked_cats:
                self.fil_mat.addItem(m["nombre"], m["id"])

        # Intentar restaurar selecciones previas de material
        # (Esto es complejo con CheckableComboBox, lo saltamos por simplicidad UX, 
        # el usuario tendrá que re-seleccionar material si cambia categoría).

        # --- Repoblar MEDIDA ---
        self.fil_medida.clear()
        
        # Aquí es complejo vincular medida a categoría sin datos del backend.
        # Heurística: Si solo LONA chequeado, mostrar medidas típicas lona (>200).
        # Si ADHESIVO chequeado, mostrar típicas (<160). 
        # Si AMBOS o NINGUNO, mostrar todas.
        
        soloadh = 1 in checked_cats and 0 not in checked_cats
        sololona = 0 in checked_cats and 1 not in checked_cats

        for w in self._all_widths_list:
            if show_all:
                self.fil_medida.addItem(f"{w:.1f}", w)
            elif soloadh and w < 160: # Heurística medidas adhesivo
                self.fil_medida.addItem(f"{w:.1f}", w)
            elif sololona and w > 160: # Heurística medidas lona
                self.fil_medida.addItem(f"{w:.1f}", w)
            elif not soloadh and not sololona: # Ambos chequeados
                self.fil_medida.addItem(f"{w:.1f}", w)


    def _add_row(self, r: dict):
        i = self.tbl.rowCount()
        self.tbl.insertRow(i)

        def set_text(c: int, val, tip: str | None = None):
            it = QTableWidgetItem("" if val is None else str(val))
            if tip:
                it.setToolTip(tip)
            self.tbl.setItem(i, c, it)

        # ----- 0: ID (Guardamos Categoría oculta) -----
        mat_nombre = r.get("material", "")
        # Heurística: 1 para Adhesivo, 0 para Lona
        cat_id = 1 if self._es_adhesivo_nombre(mat_nombre) else 0
        
        it_id = QTableWidgetItem(str(r.get("id_orden", "")))
        it_id.setData(Qt.UserRole, cat_id) # 💾 GUARDADO OCULTO DE CATEGORÍA
        self.tbl.setItem(i, 0, it_id)

        # ----- 1..3: Estado y Fechas -----
        set_text(1, r.get("estado", ""))
        set_text(2, r.get("FECHA", ""))

        # Entrega (col 3): la seteamos normal y luego la estilamos abajo
        fecha_entrega_str = str(r.get("fecha_entrega", "") or "")
        set_text(3, fecha_entrega_str)

        # ----- 4: Material (Guardamos Material ID oculto) -----
        # Si el backend no envía el ID del material, lo buscamos en el catálogo maestro
        mat_id = r.get("material_id")
        if mat_id is None:
            for m in getattr(self, "_all_materials_list", []):
                if m.get("nombre") == mat_nombre:
                    mat_id = m.get("id")
                    break

        it_mat = QTableWidgetItem(mat_nombre)
        it_mat.setData(Qt.UserRole, mat_id) # 💾 GUARDADO OCULTO DE MATERIAL ID
        self.tbl.setItem(i, 4, it_mat)

        # ----- 5..7: Ancho, Largo, Reps (texto plano) -----
        set_text(5, r.get("ancho_rollo_cm", ""))
        set_text(6, r.get("LARGO", ""))
        set_text(7, r.get("REP", ""))

        # ----- 8: Desp.L -----
        sb_desp_l = QDoubleSpinBox()
        sb_desp_l.setDecimals(2)
        sb_desp_l.setMaximum(10_000_000)
        sb_desp_l.setValue(0.0)
        sb_desp_l.valueChanged.connect(lambda _v, row=i: self._update_consumo_cell(row))
        self.tbl.setCellWidget(i, 8, sb_desp_l)


        # ----- 9: Desp.A -----
        try:
            ancho_rollo = float(r.get("ancho_rollo_cm")) if r.get("ancho_rollo_cm") is not None else None
        except Exception:
            ancho_rollo = None

        ancho_arte = None
        for k in ("ancho_orden_cm", "ANCHO_ORDEN_CM", "ancho_arte_cm", "ANCHO_ARTE_CM"):
            if r.get(k) is not None:
                try:
                    ancho_arte = float(r.get(k)); break
                except Exception:
                    pass

        sb_desp_a = QDoubleSpinBox()
        sb_desp_a.setDecimals(2)
        sb_desp_a.setMaximum(10_000_000)

        if (ancho_rollo is not None) and (ancho_arte is not None):
            sb_desp_a.setValue(max(0.0, ancho_rollo - ancho_arte))
            sb_desp_a.setEnabled(False)
            sb_desp_a.setToolTip("Calculado automáticamente: ancho del rollo - ancho del arte")
        else:
            sb_desp_a.setValue(0.0)
            sb_desp_a.setToolTip("No se pudo calcular; puedes ingresarlo manualmente.")
        self.tbl.setCellWidget(i, 9, sb_desp_a)

        # ----- 10..12: Laminado -----
        sin_lam_id = self._id_sin_laminado()

        cb_tipo = QComboBox()
        for t in self._lam_tipos:
            cb_tipo.addItem((t.get("nombre") or "").strip(), t.get("id"))
        self.tbl.setCellWidget(i, 10, cb_tipo)

        lam_tipo_id  = r.get("lam_tipo_id")  or r.get("LAM_TIPO_ID")
        lam_tipo_nom = (r.get("lam_tipo")    or r.get("LAM_TIPO") or "").strip()
        sel = -1
        if lam_tipo_id is not None:
            sel = cb_tipo.findData(lam_tipo_id)
        if sel < 0 and lam_tipo_nom:
            for idx in range(cb_tipo.count()):
                if (cb_tipo.itemText(idx) or "").strip().upper() == lam_tipo_nom.upper():
                    sel = idx; break
        if sel < 0 and sin_lam_id is not None:
            sel = cb_tipo.findData(sin_lam_id)
        if sel < 0 and cb_tipo.count() > 0:
            sel = 0
        cb_tipo.setCurrentIndex(sel)
        cb_tipo.setEnabled(False)  # lo define Diseño

        cb_marca = QComboBox()
        cb_marca.addItem("—", None)
        
        # 1. Variable para rastrear la posición de ARCLAD si existe
        indice_arclad = -1 
        
        for m in self._lam_marcas:
            cb_marca.addItem(m["nombre"], m["id"])
            # Buscamos "ARCLAD" de forma segura ignorando mayúsculas/minúsculas
            if "ARCLAD" in str(m["nombre"]).upper():
                indice_arclad = cb_marca.count() - 1
                
        self.tbl.setCellWidget(i, 11, cb_marca)

        lam_marca_id = r.get("lam_marca_id") or r.get("LAM_MARCA_ID")
        marca_asignada = False
        
        if lam_marca_id is not None:
            for idx in range(cb_marca.count()):
                if cb_marca.itemData(idx) == lam_marca_id:
                    cb_marca.setCurrentIndex(idx)
                    marca_asignada = True
                    break
        
        # 2. Asignación automática segura
        # Si no venía ninguna marca de la BD, y ARCLAD existe en el catálogo, se asigna.
        if not marca_asignada and indice_arclad != -1:
            cb_marca.setCurrentIndex(indice_arclad)

        cb_ancho = QComboBox()
        for a in self._lam_medidas:
            cb_ancho.addItem(str(a["ancho"]), float(a["ancho"]))
        self.tbl.setCellWidget(i, 12, cb_ancho)

        material_nombre = r.get("material") or ""
        es_adh = (bool(r.get("es_adhesivo"))
                if "es_adhesivo" in r else self._es_adhesivo_nombre(material_nombre))

        try:
            ancho_rollo = float(r.get("ancho_rollo_cm")) if r.get("ancho_rollo_cm") is not None else None
        except Exception:
            ancho_rollo = None

        es_sin_lam = ((cb_tipo.currentText() or "").strip().upper() == "SIN LAMINAR"
                    or (sin_lam_id is not None and cb_tipo.currentData() == sin_lam_id))

        if es_sin_lam:
            cb_marca.setCurrentIndex(0); cb_marca.setEnabled(False)
            cb_ancho.clear(); cb_ancho.addItem("0.0", 0.0); cb_ancho.setCurrentIndex(0); cb_ancho.setEnabled(False)
        else:
            if ancho_rollo is not None:
                found = False
                for idx in range(cb_ancho.count()):
                    try:
                        if float(cb_ancho.itemData(idx)) == float(ancho_rollo):
                            cb_ancho.setCurrentIndex(idx); found = True; break
                    except Exception:
                        pass
                if not found:
                    cb_ancho.insertItem(0, str(ancho_rollo), float(ancho_rollo))
                    cb_ancho.setCurrentIndex(0)
            cb_ancho.setEnabled(not es_adh)

        # ----- 13: Consumo lam / 14: Merma -----
        sb_lam_cons = QDoubleSpinBox()
        sb_lam_cons.setDecimals(2)
        sb_lam_cons.setMaximum(10_000_000)
        sb_lam_cons.setValue(float(r.get("lam_consumo_cm") or r.get("LAM_CONSUMO_CM") or 0.0))

        if es_sin_lam:
            sb_lam_cons.setValue(float(r.get("lam_consumo_cm") or r.get("LAM_CONSUMO_CM") or 0.0))
            sb_lam_cons.setEnabled(False)
            sb_lam_cons.setToolTip("SIN LAMINAR: consumo = 0.00 (no editable)")
        self.tbl.setCellWidget(i, 13, sb_lam_cons)

        sb_merma = QDoubleSpinBox()
        sb_merma.setDecimals(2)
        sb_merma.setMaximum(10_000_000)
        sb_merma.setValue(0.0)
        self.tbl.setCellWidget(i, 14, sb_merma)

        # ----- 15: Esp.reps (guarda al vuelo) -----
        sb_esp = QDoubleSpinBox()
        sb_esp.setDecimals(2)
        sb_esp.setMaximum(10_000_000)
        sb_esp.setValue(float(r.get("espacio_reps_cm") or 0.0))
        oid = int(r.get("id_orden", 0) or 0)
        sb_esp.valueChanged.connect(lambda v, _oid=oid: self._guardar_esp_reps(_oid, v))
        sb_esp.valueChanged.connect(lambda _v, row=i: self._update_consumo_cell(row))

        self.tbl.setCellWidget(i, 15, sb_esp)

        # ----- 16: Observaciones (solo lectura) -----
        obs_txt = str(r.get("observaciones") or "")
        it_obs = QTableWidgetItem(obs_txt)
        it_obs.setToolTip(obs_txt)
        it_obs.setFlags(it_obs.flags() & ~Qt.ItemIsEditable)
        self.tbl.setItem(i, 16, it_obs)

        # ----- 17: Ruta -----
        ruta_local = r.get("ruta") or r.get("RUTA") or ""
        set_text(17, ruta_local)
        it_ruta = self.tbl.item(i, 17)
        if it_ruta:
            it_ruta.setToolTip(ruta_local)


        it_cons = QTableWidgetItem("")
        it_cons.setFlags(it_cons.flags() & ~Qt.ItemIsEditable)
        self.tbl.setItem(i, 18, it_cons)
        self._update_consumo_cell(i)

        # ----- 18: Acciones -----
        cell = QWidget()
        h = QHBoxLayout(cell); h.setContentsMargins(6, 4, 6, 4); h.setSpacing(8)

        b_open = QPushButton("Abrir")
        b_ep   = QPushButton("En pr")
        b_fi   = QPushButton("Final")
        b_ca   = QPushButton("Canc")

        for b in (b_open, b_ep, b_fi, b_ca):
            b.setMinimumHeight(28)
            b.setCursor(Qt.PointingHandCursor)

        h.addWidget(b_open); h.addWidget(b_ep); h.addWidget(b_fi); h.addWidget(b_ca); h.addStretch()
        self.tbl.setCellWidget(i, 19,cell)

        ruta_txt = r.get("ruta", "")
        b_open.clicked.connect(lambda _=False, p=ruta_txt: self._abrir_ruta(p))

        oid = int(r.get("id_orden", 0) or 0)
        b_ep.clicked.connect(lambda _=False, row=i, _oid=oid: self._row_en_proceso(row, _oid))
        b_fi.clicked.connect(lambda _=False, row=i, _oid=oid: self._row_finalizar(row, _oid))
        b_ca.clicked.connect(lambda _=False, row=i, _oid=oid: self._row_cancelar(row, _oid))

        # --- Colores de fila por estado (igual que ya tenías) ---
        estado_txt = r.get("estado", "")
        est = (estado_txt or "").upper()
        if   "FINAL"   in est: self._paint_row_bg(i, FINALIZADA_BG)
        elif "CANCEL"  in est: self._paint_row_bg(i, CANCELADA_BG)
        elif "PROCESO" in est: self._paint_row_bg(i, EN_PROCESO_BG)
        else:                  self._paint_row_bg(i, PENDIENTE_BG)

        # ✅ NUEVO: aplicar emoji / naranja SOLO en celda Entrega
        self._style_entrega_cell(i, fecha_entrega_str)



    #Detectar si el material es un adhesivo:
    def _es_adhesivo_nombre(self, nombre: str | None) -> bool:
        """Heurística: detecta adhesivos por texto del material.
        Sugerencia: si puedes, haz que el backend envíe un flag 'es_adhesivo' o 'categoria'."""
        n = (nombre or "").strip().upper()
        needles = ("ADH", "ADHESIVO", "VINIL", "VINILO", "STICKER")
        return any(tok in n for tok in needles)    

        

    # ---------- helpers de widgets ---------
    def _select_combo(self, combo: QComboBox, value):
        for idx in range(combo.count()):
            try:
                if combo.itemData(idx) == value:
                    combo.setCurrentIndex(idx); return
            except Exception:
                pass

    def _put_dspin(self, row, col, val):
        sb = QDoubleSpinBox(); sb.setDecimals(2); sb.setMaximum(10_000_000)
        sb.setValue(float(val or 0))
        self.tbl.setCellWidget(row, col, sb)
        return sb

    def _put_spin(self, row, col, val):
        sp = QSpinBox(); sp.setMaximum(10000); sp.setValue(int(val or 0))
        self.tbl.setCellWidget(row, col, sp)
        return sp

    def _put_lam_tipo(self, row, col, preselect_id=None):
        cb = QComboBox()
        for t in self._lam_tipos:
            cb.addItem(t["nombre"], t["id"])
        if preselect_id is not None:
            self._select_combo(cb, preselect_id)
        self.tbl.setCellWidget(row, col, cb)
        return cb

    def _put_lam_marca(self, row, col, preselect_id=None):
        cb = QComboBox()
        cb.addItem("--", None)  # opción nula
        for m in self._lam_marcas:
            cb.addItem(m["nombre"], m["id"])
        if preselect_id is not None:
            self._select_combo(cb, preselect_id)
        self.tbl.setCellWidget(row, col, cb)
        return cb

    def _put_lam_ancho(self, row, col, preselect_val=None):
        cb = QComboBox()
        for a in self._lam_medidas:
            cb.addItem(str(a["ancho"]), float(a["ancho"]))
        if preselect_val is not None:
            for idx in range(cb.count()):
                try:
                    if float(cb.itemData(idx)) == float(preselect_val):
                        cb.setCurrentIndex(idx); break
                except Exception:
                    pass
        self.tbl.setCellWidget(row, col, cb)
        return cb

    # ---------- getter de valores ----------
    def _g(self, row, col, cast=float):
        w = self.tbl.cellWidget(row, col)
        if isinstance(w, QDoubleSpinBox):
            return float(w.value())
        if isinstance(w, QSpinBox):
            return int(w.value())
        if isinstance(w, QComboBox):
            d = w.currentData()
            try:
                return cast(d)
            except Exception:
                return d
        it = self.tbl.item(row, col)
        return cast(it.text()) if it else None
    

    # ---------- validación ----------
    def _validate_finalize(self, row) -> bool:
        """Valida que se pueda FINALIZAR."""
        missing = []

        # --- Desp.L > 0 obligatorio ---
        desp_l = self._g(row, 8, float)
        if desp_l is None or desp_l <= 0:
            missing.append("Desp.L(cm) (> 0)")

        # Material (heurística adhesivo)
        mat_item = self.tbl.item(row, 4)
        material_nombre = mat_item.text() if mat_item else ""
        es_adh = self._es_adhesivo_nombre(material_nombre)

        # Datos comunes de fila
        w_tipo = self.tbl.cellWidget(row, 10)
        lam_tipo_txt = (w_tipo.currentText() if isinstance(w_tipo, QComboBox) else "").strip().upper()
        lam_ancho = self._g(row, self.COL_LAM_ANCHO, float) or 0.0
        lam_cons  = self._g(row, self.COL_LAM_CONS, float)
        lam_marca = self._g(row, self.COL_LAM_MARCA, int)

        rep = self._g(row, 7, int) or 0
        largo = self._g(row, 6, float) or 0.0
        esp = self._g(row, 15, float) or 0.0

        # Consumo total esperado en IMPRESIÓN
        esperado = max(0.0, largo * max(rep, 0) + (desp_l or 0.0) + esp * max(rep, 0))

        # --- Reglas por tipo de laminado ---
        if lam_tipo_txt != "SIN LAMINAR":
            # Marca requerida
            if lam_marca is None:
                missing.append("Lam.marca")
            # Ancho > 0
            if lam_ancho <= 0:
                missing.append("Lam.ancho(cm) (>0)")

            # Consumo obligatorio y >= esperado
            if lam_cons is None or lam_cons <= 0:
                missing.append("Lam.consumo(cm) (obligatorio > 0)")
            else:
                if float(lam_cons) + 1e-6 < float(esperado):
                    missing.append(f"Lam.consumo(cm) (≥ consumo esperado: {esperado:.2f})")

            # Adhesivo ⇒ lam.ancho == ancho rollo
            try:
                ancho_roll = float(self.tbl.item(row, self.COL_ANCHO_ROLL).text())

            except Exception:
                ancho_roll = None
            if es_adh and ancho_roll is not None and abs(lam_ancho - ancho_roll) > 0.01:
                missing.append("Lam.ancho debe ser igual a 'ancho roll' para adhesivos")

        else:
            # SIN LAMINAR: consumo fijo 0.00
            if (lam_cons or 0.0) != 0.0:
                missing.append("Lam.consumo debe ser 0.00 cuando Lam.tipo = SIN LAMINAR")

        # REP ≥ 1 ⇒ Esp.reps > 0 obligatorio
        if rep >= 1 and (esp is None or esp <= 0):
            missing.append("Esp.reps(cm) (>0) es obligatorio cuando REP ≥ 1")

        if missing:
            QMessageBox.warning(self, "No se puede finalizar", "Faltan/ajusta:\n- " + "\n- ".join(missing))
            return False
        return True



    # ---------- acciones ----------
    def _row_en_proceso(self, row, oid):
        ancho_obj = self._g(row, 12, float)  # Lam.ancho
        payload = {"id_orden": oid, "usuario": None, "ancho_objetivo_cm": ancho_obj}
        r = api_post("/ordenes/en-proceso", payload, self)
        if r:
            self._load_panel()

    def _row_finalizar(self, row, oid):
        if not self._validate_finalize(row):
            return

        tipo_id = self._g(row, self.COL_LAM_TIPO, int)
        if tipo_id is None:
            tipo_id = self._id_sin_laminado()
            if tipo_id is None:
                QMessageBox.warning(self, "Error", "No encuentro el ID de 'SIN LAMINAR' en el catálogo. No se puede finalizar.")
                return

        # ✅ 1) Calcula variables ANTES del dict
        lam_ancho = self._g(row, self.COL_LAM_ANCHO, float)
        lam_cons  = self._g(row, self.COL_LAM_CONS, float)

        payload = {
            "id_orden": oid,
            "total_largo_impreso_cm": self._g(row, self.COL_LARGO, float),
            "desp_largo_cm":           self._g(row, self.COL_DESP_L, float),
            "desp_ancho_cm":           self._g(row, self.COL_DESP_A, float),
            "repeticiones":            self._g(row, self.COL_REP, int),
            "usuario_impresor": None,
            "estado_requerido": "EN PROCESO",
            "ancho_objetivo_cm":       lam_ancho,

            "lam_tipo_id":    tipo_id,
            "lam_marca_id":   self._g(row, self.COL_LAM_MARCA, int),
            "lam_ancho_cm":   lam_ancho,
            "lam_consumo_cm": lam_cons,

            # ✅ COMPAT (por si tu backend usa otras claves)
            "lam_ancho":      lam_ancho,
            "lam_consumo":    lam_cons,
            "LAM_ANCHO_CM":   lam_ancho,
            "LAM_CONSUMO_CM": lam_cons,
        }


        r = api_post("/ordenes/finalizar", payload, self)
        if r:
            self._load_panel()


    def _row_cancelar(self, row, oid):
        merma = self._g(row, 14, float)
        if merma is None or merma <= 0:
            QMessageBox.warning(self, "No se puede cancelar", "Debes diligenciar Merma cancel(cm) (>0).")
            return
        payload = {
            "id_orden": oid,
            # Enviamos ambos nombres por compatibilidad con el backend actual
            "desp_largo_cm": merma,
            "merma_cancel_cm": merma,
            "usuario": None,
            "estado_permitido": ["PENDIENTE","EN PROCESO"],
            "ancho_objetivo_cm": self._g(row, 12, float),
            "volver_a_pendiente": True
        }
        r = api_post("/ordenes/cancelar", payload, self)
        if r:
            self._load_panel()

    def _limpiar_finalizadas(self):
        r = api_post("/ordenes/limpiar-finalizadas", {}, self)
        if r:
            QMessageBox.information(self, "OK", f"Archivadas: {r.get('archivadas',0)}")
            self._load_panel()

    # ==========================================
    # LÓGICA DE PALETA DE COLORES (Clic Derecho)
    # ==========================================
    # ========================================================
    # LÓGICA AVANZADA DE PALETA DE COLORES (Clic Derecho)
    # ========================================================
    def _show_context_menu(self, pos):
        """Muestra un menú profesional al hacer clic derecho en una fila."""
        # 1. Identificar fila
        item = self.tbl.itemAt(pos)
        if not item: return
        row = item.row()
        #Considerar backgraund de la paleta de colores frente a los cambios implementados
        # 2. Crear menú base
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: white; border: 1px solid #ccc; padding: 5px; }
            QMenu::item { padding: 8px 25px; border-radius: 4px; }
            QMenu::item:selected { background-color: #e3f2fd; color: black; }
            QMenu::separator { height: 1px; background: #e0e0e0; margin: 5px 0; }
        """)
        
        # --- SECCIÓN 1: COLORES PREDEFINIDOS (Rápido) ---
        paleta_fija = [
            ("🔴 Rojo ", "#ffcdd2"),

            ("🟡 Amarillo ", "#fff9c4"),

            ("🟢 Verde ", "#c8e6c9"),

            ("🔵 Azul ", "#bbdefb"),

            ("🟣 Violeta ", "#5E2266") , 

            ("   Aguamarina " , "#13DD6E")
        ]
        
        for nombre, hex_color in paleta_fija:
            accion = QAction(self._crear_icono_color(hex_color), nombre, self)
            accion.triggered.connect(lambda checked=False, r=row, c=hex_color: self._apply_color_to_row(r, c))
            menu.addAction(accion)

        menu.addSeparator()
        
        # --- SECCIÓN 2: COLORES RECIENTES (Herramienta de reutilización) ---
        if self._colores_recientes:
            menu_recientes = menu.addMenu("🔄 Reutilizar Colores Recientes")
            for hex_color in self._colores_recientes:
                # Usamos el código hexadecimal como nombre del item
                accion = QAction(self._crear_icono_color(hex_color), hex_color, self)
                accion.triggered.connect(lambda checked=False, r=row, c=hex_color: self._apply_color_to_row(r, c))
                menu_recientes.addAction(accion)
            menu.addSeparator()

        # --- SECCIÓN 3: COLOR PERSONALIZADO (Paleta Completa) ---
        accion_custom = QAction(QIcon(), "🎨 Elegir Color Personalizado...", self)
        # Conectamos a una nueva función que abre el diálogo
        accion_custom.triggered.connect(lambda checked=False, r=row: self._pick_custom_color(r))
        menu.addAction(accion_custom)

        menu.addSeparator()

        # --- SECCIÓN 4: QUITAR COLOR ---
        accion_reset = QAction("⚪ Quitar Color (Restaurar)", self)
        accion_reset.triggered.connect(lambda checked=False, r=row: self._apply_color_to_row(r, "#ffffff"))
        menu.addAction(accion_reset)

        # Mostrar menú
        menu.exec_(self.tbl.viewport().mapToGlobal(pos))


    # --- HELPERS PARA LA LÓGICA DE COLOR ---

    def _crear_icono_color(self, hex_color: str) -> QIcon:
        """Helper para crear un icono cuadrado pequeño del color indicado."""
        pixmap = QPixmap(16, 16) # Tamaño del icono
        pixmap.fill(QColor(hex_color))
        return QIcon(pixmap)

    def _pick_custom_color(self, row: int):
        """Abre el diálogo de QColorDialog y gestiona el guardado."""
        # 1. Obtener color actual de la fila como punto de partida (opcional, pero buena UX)
        current_item = self.tbl.item(row, 0)
        current_color = current_item.background().color() if current_item else Qt.white

        # 2. Abrir QColorDialog
        color = QColorDialog.getColor(current_color, self, "Seleccionar Color Personalizado")
        
        if color.isValid():
            hex_color = color.name() # Obtenemos el código #RRGGBB
            
            # 3. Aplicar color a la fila
            self._apply_color_to_row(row, hex_color)
            
            # 4. Guardar en la lista de recientes local (QSettings) 👇
            self._add_color_to_recents(hex_color)


    def _add_color_to_recents(self, hex_color: str):
        """Añade el color a la lista de recientes, la ordena y la guarda localmente."""
        # Eliminar si ya existía (para moverlo al principio)
        if hex_color in self._colores_recientes:
            self._colores_recientes.remove(hex_color)
            
        # Añadir al principio de la lista
        self._colores_recientes.insert(0, hex_color)
        
        # Limitar a los últimos 10
        self._colores_recientes = self._colores_recientes[:10]
        
        # ✅ GUARDAR PERMANENTEMENTE EN LA COMPUTADORA DEL USUARIO 👇
        self.settings.setValue("impresion/colores_recientes", self._colores_recientes)


    def _apply_color_to_row(self, row: int, hex_color: str):
        """Aplica el color de fondo a todas las celdas de una fila (Mejorado)."""
        color = QColor(hex_color)
        
        # Bloquear actualizaciones temporalmente para velocidad visual
        self.tbl.setUpdatesEnabled(False)
        try:
            for col in range(self.tbl.columnCount()):
                item = self.tbl.item(row, col)
                if item:
                    item.setBackground(color)
                    
                    # Pequeño ajuste de UX: si el fondo es muy oscuro, poner el texto blanco
                    # (Algoritmo simple de luminosidad)
                    if color.lightness() < 128:
                        item.setForeground(Qt.white)
                    else:
                        item.setForeground(Qt.black) # O restaurar color original
        finally:
            self.tbl.setUpdatesEnabled(True)

# =========================
#          MAIN
# =========================
class Main(QMainWindow):
    def _wrap_scroll(self, widget: QWidget) -> QScrollArea:
        sa = QScrollArea()
        sa.setWidget(widget)
        sa.setWidgetResizable(True)
        sa.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        sa.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        return sa

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Control de Producción - Gran Formato")

        # Icono de la ventana (opcional)
        app_icon_path = os.path.join(os.path.dirname(__file__), "assets", "icons", "app_icon.png")
        if os.path.exists(app_icon_path):
            self.setWindowIcon(QIcon(app_icon_path))

        # === 1) Login ===
        dlg = DlgLogin(self)
        if dlg.exec() != QDialog.Accepted or not dlg.result_data:
            sys.exit(0)

        auth = dlg.result_data or {}
        token   = auth.get("access_token") or auth.get("token")
        rol_raw = (auth.get("rol") or "").upper()

        if not token:
            QMessageBox.critical(self, "Login", "No se recibió access_token del servidor.")
            sys.exit(1)

        # Si el subusuario quedó pendiente/rechazado, no permitimos entrar
        if rol_raw in {"PENDIENTE", "RECHAZADO"}:
            QMessageBox.warning(
                self,
                "Acceso pendiente",
                "Tu subusuario está PENDIENTE (o fue RECHAZADO).\n"
                "Pídele a un ADMIN que te asigne un rol (DISEÑO / IMPRESIÓN / ADMIN) y te habilite."
            )
            sys.exit(0)

        # --- Identidad: email principal + etiqueta (subusuario) ---
        email_principal = (
            auth.get("email_principal")
            or auth.get("email")
            or auth.get("usr")
            or ""
        )
        subusuario = (
            auth.get("usuario")              # login-subuser
            or auth.get("subusuario")        # compat
            or auth.get("nombre_usuario")    # fallback legacy
            or ""
        )
        sid = auth.get("sub_usuario_id") or auth.get("sid") or ""

        self.setWindowTitle(
            f"Control de Producción - Gran Formato · {email_principal} · {subusuario} · {rol_raw or '—'}"
        )

        # 2) Asegura que el statusbar esté visible
        self.statusBar().setVisible(True)

        # 3) Widget permanente (no se borra con showMessage)
        session_txt = f"Sesión: {email_principal} · Subusuario: {subusuario} · Rol: {rol_raw or '—'}"
        self._lab_session = QLabel(session_txt)
        self._lab_session.setStyleSheet("padding: 0 8px;")
        self.statusBar().addPermanentWidget(self._lab_session, 1)


        # === 2) Estado de auth global ===
        set_token(token)
        global AUTH_TOKEN, API_TOKEN, USER_ROLES
        AUTH_TOKEN = token
        API_TOKEN  = token
        USER_ROLES = [rol_raw] if rol_raw else []

        # Barra de estado: sesión PERMANENTE + mensajes temporales (WS, etc.)
        session_txt = f"Sesión: {email_principal} · Subusuario: {subusuario} · Rol: {rol_raw or '—'}"
        if sid != "":
            session_txt += f" · SID: {sid}"

        self._lab_session = QLabel(session_txt)
        self._lab_session.setStyleSheet("color:#111; padding: 0 8px;")
        self.statusBar().addPermanentWidget(self._lab_session, 1)

        # === 3) Páginas ===
        self.pag_disenio    = PagDisenio()
        self.pag_impresion  = PagImpresion()
        self.pag_inventario = PagInventario()
        self.pag_reportes   = PagReportes()
        self.pag_admin      = PagAdmin(API, AUTH_TOKEN)

        # === 4) Sidebar + páginas según rol ===
        roles = {r.upper() for r in (USER_ROLES or [])}
        if "ADMIN" in roles:
            allowed = [
                ("Diseño",         self.pag_disenio),
                ("Impresión",      self.pag_impresion),
                ("Inventario",     self.pag_inventario),
                ("Reportes",       self.pag_reportes),
                ("Administración", self.pag_admin),
            ]
            default_index = 1
        elif "DISENO" in roles:
            allowed = [
                ("Diseño",         self.pag_disenio),
                ("Impresión",      self.pag_impresion),
                ("Inventario",     self.pag_inventario),
                ("Reportes",       self.pag_reportes),
            ]
            default_index = 1
        elif "IMPRESION" in roles:
            allowed = [
                ("Impresión",  self.pag_impresion),
                ("Inventario", self.pag_inventario),
                ("Reportes",   self.pag_reportes),
            ]
            default_index = 0
        else:
            allowed = [
                ("Impresión",  self.pag_impresion),
                ("Inventario", self.pag_inventario),
                ("Reportes",   self.pag_reportes),
            ]
            default_index = 0

        # --- contenedor izquierdo (logo + lista) ---
        left_panel = QWidget()
        left_v = QVBoxLayout(left_panel)
        left_v.setContentsMargins(8, 8, 8, 8)
        left_v.setSpacing(8)

        # Botón del logo (toggle)
        self.btn_logo = QPushButton()
        self.btn_logo.setFlat(True)
        self.btn_logo.setCursor(Qt.PointingHandCursor)
        self.btn_logo.setIcon(_icon_or_fallback(LOGO_PATH, left_panel, QStyle.SP_DesktopIcon))
        self.btn_logo.setToolTip("Ocultar menú")
        self.btn_logo.clicked.connect(self._toggle_sidebar)

        self._collapsed_w = 44
        self._sidebar_last_w = 180
        self._sidebar_collapsed = False

        # el botón debe seguir visible y cómodo en modo colapsado
        self.btn_logo.setFixedSize(32, 32)
        self.btn_logo.setIconSize(QSize(22, 22))
        left_v.addWidget(self.btn_logo, 0, Qt.AlignHCenter)

        # Lista y páginas
        self.sidebar = QListWidget()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setIconSize(QSize(18, 18))
        self.sidebar.setUniformItemSizes(True)
        self.sidebar.setSpacing(2)
        self.sidebar.setStyleSheet("""
        #Sidebar {
            background: #fafafa;
            border: 1px solid #e5e5e5;
            border-radius: 10px;
        }
        #Sidebar::item {
            padding: 8px 12px;
            border-radius: 10px;
            margin: 2px 6px;
        }
        #Sidebar::item:selected {
            background: #e8f0fe;
            color: #0b57d0;
        }
        #Sidebar::item:hover {
            background: #f2f6ff;
        }
        """)

        self.pages = QStackedWidget()

        # Íconos de la lista — firma correcta: (path, widget, fallback_sp)
        base_ic = os.path.join(os.path.dirname(__file__), "assets", "icons")
        icon_map = {
            "Diseño":         _icon_or_fallback(os.path.join(base_ic, "design.svg"),     self, QStyle.SP_FileDialogDetailedView),
            "Impresión":      _icon_or_fallback(os.path.join(base_ic, "print.svg"),      self, QStyle.SP_DesktopIcon),
            "Inventario":     _icon_or_fallback(os.path.join(base_ic, "inventory.svg"),  self, QStyle.SP_DirIcon),
            "Reportes":       _icon_or_fallback(os.path.join(base_ic, "report.svg"),     self, QStyle.SP_FileIcon),
            "Administración": _icon_or_fallback(os.path.join(base_ic, "admin.svg"),      self, QStyle.SP_ComputerIcon),
        }

        for name, widget in allowed:
            icon = icon_map.get(name, self.style().standardIcon(QStyle.SP_FileIcon))
            self.sidebar.addItem(_make_sidebar_item(name, icon))
            self.pages.addWidget(self._wrap_scroll(widget))

        self.sidebar.currentRowChanged.connect(self.pages.setCurrentIndex)
        self.sidebar.setCurrentRow(default_index)

        left_v.addWidget(self.sidebar, 1)

        # === 5) Splitter principal ===
        self.splitter = QSplitter()
        self.splitter.addWidget(left_panel)
        self.splitter.addWidget(self.pages)
        self.splitter.setCollapsible(0, True)
        self.splitter.setStretchFactor(0, 0)
        self.splitter.setStretchFactor(1, 1)

        # Permitir resize manual del menú (no usar Fixed)
        left_panel.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)

        # tamaños iniciales
        self.splitter.setSizes([self._sidebar_last_w, 1100])
        self.splitter.splitterMoved.connect(self._on_splitter_moved)

        self.setCentralWidget(self.splitter)
        self._left_panel = left_panel

        # Fondo (opcional)
        try:
            if APP_BG_PATH and APP_BG_PATH.exists():
                bg_url = APP_BG_PATH.as_posix()
                self.setStyleSheet(
                    self.styleSheet() +
                    f"""
                    QMainWindow {{
                        background-image: url("{bg_url}");
                        background-position: center;
                        background-repeat: no-repeat;
                        background-attachment: fixed;
                        background-color: #111111;
                    }}
                    """
                )
        except Exception:
            pass

        # === 6) WebSocket ===
        base_ws_url = _ws_url_from_api(API)
        ws_url = f"{base_ws_url}?token={API_TOKEN}" if API_TOKEN else base_ws_url
        self.ws = WSClient(ws_url, self)
        bus = self.ws.bus

        bus.orderCreated.connect(lambda _p: self.pag_impresion._load_panel())
        bus.orderFinalized.connect(lambda _p: (self.pag_impresion._load_panel(), self.pag_reportes._load()))
        bus.orderCanceled.connect(lambda _p: (self.pag_impresion._load_panel(), self.pag_reportes._load()))
        bus.connected.connect(lambda: self.statusBar().showMessage("WS conectado", 3000))
        bus.disconnected.connect(lambda why: self.statusBar().showMessage(f"WS desconectado: {why}", 3000))
        self.ws.connect()

    # --- recordar ancho al mover el divisor ---
    def _on_splitter_moved(self, pos: int, index: int):
        if not self._sidebar_collapsed:
            w = self._left_panel.width()
            if w >= 120:
                self._sidebar_last_w = w

    # --- plegar / desplegar panel izquierdo ---
    def _toggle_sidebar(self):
        self._sidebar_collapsed = not self._sidebar_collapsed

        if self._sidebar_collapsed:
            # Guardar el ancho actual para volver
            self._sidebar_last_w = max(120, self._left_panel.width())

            # Deja el panel angosto pero visible
            self._left_panel.setMinimumWidth(self._collapsed_w)
            self._left_panel.setMaximumWidth(self._collapsed_w)
            self.sidebar.setVisible(False)

            lp_layout: QVBoxLayout = self._left_panel.layout()
            if lp_layout:
                lp_layout.setContentsMargins(4, 8, 4, 8)

            self.splitter.setSizes([self._collapsed_w, max(600, self.width() - self._collapsed_w)])
            self.btn_logo.setToolTip("Mostrar menú")

        else:
            w = max(120, self._sidebar_last_w)

            self._left_panel.setMaximumWidth(16777215)
            self._left_panel.setMinimumWidth(w)
            self.sidebar.setVisible(True)

            lp_layout: QVBoxLayout = self._left_panel.layout()
            if lp_layout:
                lp_layout.setContentsMargins(8, 8, 8, 8)

            self.splitter.setSizes([w, max(600, self.width() - w)])
            self.btn_logo.setToolTip("Ocultar menú")




# --- Lanzador de la app de escritorio ---
if __name__ == "__main__":
    import ctypes
    _dbg("=== Desktop arrancando (stdout line-buffered) ===")
    if sys.platform.startswith("win"):
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("GF.ControlProduccion")
        except Exception:
            pass

    qtapp = QApplication(sys.argv)
    qtapp.setApplicationName("Control de Producción - Gran Formato")
    qtapp.setWindowIcon(_icon_or_fallback(LOGO_PATH, QWidget(), QStyle.SP_DesktopIcon))
    apply_theme(qtapp, mode="light")  

    win = Main()
    win.resize(1300, 800)
    win.show()
    sys.exit(qtapp.exec())



#input
if __name__ == "__main__":
    import ctypes
    _dbg("=== Desktop arrancando (stdout line-buffered) ===")
    if sys.platform.startswith("win"):
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("GF.ControlProduccion")
        except Exception:
            pass

    qtapp = QApplication(sys.argv)
    qtapp.setApplicationName("Control de Producción - Gran Formato")
    qtapp.setWindowIcon(_icon_or_fallback(LOGO_PATH, QWidget(), QStyle.SP_DesktopIcon))
    apply_theme(qtapp, mode="light")  

    win = Main()
    win.resize(1300, 800)
    win.show()
    sys.exit(qtapp.exec())

    input("Presiona Enter para salir...")  # Esta línea mantiene la consola abierta
